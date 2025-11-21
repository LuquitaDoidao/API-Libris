import os
from io import BytesIO
from flask import Flask, jsonify, request, send_file, url_for, current_app, render_template, render_template_string
from main import app, con  # Importando o app e a conexão com o banco
from flask_bcrypt import generate_password_hash, check_password_hash
from fpdf import FPDF
import jwt
import json
from datetime import datetime, time, timedelta
from flask_mail import Mail, Message
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.header import Header
import smtplib
import shutil
import qrcode
import email.charset
from qrcode.constants import ERROR_CORRECT_H
import crcmod
import requests
import random

import locale  # Importe o módulo 'locale' no topo do seu arquivo Python

try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    print("AVISO: Locale 'pt_BR.UTF-8' não encontrado. Usando o padrão do sistema.")
    # Em alguns sistemas (como Windows), você pode precisar usar 'Portuguese_Brazil.1252' ou 'pt-BR'

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= IMAGENS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= TOKEN =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
app.config.from_pyfile('config.py')
senha_secreta = app.config['SECRET_KEY']


def generate_token(user_id):
    payload = {'id_usuario': user_id}
    token = jwt.encode(payload, senha_secreta, algorithm='HS256')
    return token


def remover_bearer(token):
    if token.startswith('Bearer '):
        return token[len('Bearer '):]
    else:
        return token


def verifica_adm(id):
    cursor = con.cursor()
    cursor.execute("select id_usuario from usuarios where tipo = 3 and id_usuario = ? ", (id,))
    adm = cursor.fetchone()
    cursor.close()
    if adm:
        return True
    return False


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EMAIL =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'bibliotecalibris@gmail.com'
app.config['MAIL_PASSWORD'] = 'qahxzspwufjeyqtk'
app.config['MAIL_DEFAULT_SENDER'] = 'bibliotecalibris@gmail.com'

mail = Mail(app)


def enviar_email(destinatario, corpo_html, assunto, titulo):
    if not destinatario:
        raise ValueError("Endereço de e-mail não fornecido.")

    subject = assunto
    sender = "bibliotecalibris@gmail.com"
    recipients = [destinatario]
    password = "raij vzce iafk iekd"  # dica: use variáveis de ambiente em produção!

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = sender
        msg['To'] = ', '.join(recipients)

        # Renderiza o HTML com o template e os dados
        html_renderizado = render_template('template_email.html', corpo=corpo_html, titulo=titulo)
        # msg.attach(MIMEText(html_renderizado.encode('utf-8'), 'html', 'utf-8'))
        msg.attach(MIMEText(html_renderizado, 'html', 'utf-8'))

        # Envia o e-mail
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp_server:
            smtp_server.login(sender, password)
            smtp_server.sendmail(sender, recipients, msg.as_string())


    except Exception as e:
        print(f"Ocorreu um erro ao enviar o e-mail: {e}")


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= QRCODE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
def calcula_crc16(
        payload):  # (CRC16) a função gera esse número com base no conteúdo do payload (os dados que o QR Code vai carregar).
    crc16 = crcmod.mkCrcFun(0x11021, initCrc=0xFFFF, rev=False)
    crc = crc16(payload.encode('utf-8'))
    return f"{crc:04X}"


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PARAMETRIZAÇÃO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
RAZAO_SOCIAL = app.config['RAZAO_SOCIAL']
NOME_FANTASIA = app.config['NOME_FANTASIA']
MENSAGEM_INSTITUCIONAL = app.config['MENSAGEM_INSTITUCIONAL']
TELEFONE = app.config['TELEFONE']
EMAIL = app.config['EMAIL']
WHATSAPP = app.config['WHATSAPP']
CIDADE = app.config['CIDADE']
BAIRRO = app.config['BAIRRO']
RUA = app.config['RUA']
NUMERO = app.config['NUMERO']
PIX = app.config['PIX']
FACEBOOK = app.config['FACEBOOK']
INSTAGRAM = app.config['INSTAGRAM']
YOUTUBE = app.config['YOUTUBE']
X = app.config['X']
TIKTOK = app.config['TIKTOK']
HORARIO_FUNC_INICIO = app.config['HORARIO_FUNC_INICIO']
HORARIO_FUNC_FIM = app.config['HORARIO_FUNC_FIM']
HORARIO_ALMOCO_INICIO = app.config['HORARIO_ALMOCO_INICIO']
HORARIO_ALMOCO_FIM = app.config['HORARIO_ALMOCO_FIM']


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ADICIONA PARAMETRIZAÇÃO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/configuracoes', methods=['POST'])
def adicionar_configuracao():
    try:
        dados = request.get_json()
        if dados is None:
            return jsonify({"erro": "JSON inválido ou ausente"}), 400

        cur = con.cursor()
        cur.execute("""
            INSERT INTO CONFIGURACOES (
                RAZAO_SOCIAL, NOME_FANTASIA, MENSAGEM_INSTITUCIONAL,
                TELEFONE, EMAIL, WHATSAPP, CIDADE, BAIRRO, RUA, NUMERO, PIX,
                FACEBOOK, INSTAGRAM, YOUTUBE, X, TIKTOK,
                HORARIO_FUNCIONAMENTO_INICIO, HORARIO_FUNCIONAMENTO_FIM,
                HORARIO_ALMOCO_INICIO, HORARIO_ALMOCO_FIM
            ) VALUES (
                ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?,
                ?, ?,
                ?, ?,
            )
        """, (
            dados.get('RAZAO_SOCIAL'),
            dados.get('NOME_FANTASIA'),
            dados.get('MENSAGEM_INSTITUCIONAL'),
            dados.get('TELEFONE'),
            dados.get('EMAIL'),
            dados.get('WHATSAPP'),
            dados.get('CIDADE'),
            dados.get('BAIRRO'),
            dados.get('RUA'),
            dados.get('NUMERO'),
            dados.get('PIX'),
            dados.get('FACEBOOK'),
            dados.get('INSTAGRAM'),
            dados.get('YOUTUBE'),
            dados.get('X'),
            dados.get('TIKTOK'),
            dados.get('HORARIO_FUNCIONAMENTO_INICIO'),
            dados.get('HORARIO_FUNCIONAMENTO_FIM'),
            dados.get('HORARIO_ALMOCO_INICIO'),
            dados.get('HORARIO_ALMOCO_FIM')
        ))

        con.commit()
        cur.close()
        return jsonify({"mensagem": "Configuração adicionada com sucesso!"}), 201

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= RECUPERA PARAMETRIZAÇÃO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/configuracoes', methods=['GET'])
def obter_configuracao():
    try:

        # Retornando os dados em formato JSON
        configuracao = {
            'RAZAO_SOCIAL': RAZAO_SOCIAL,
            'NOME_FANTASIA': NOME_FANTASIA,
            'MENSAGEM_INSTITUCIONAL': MENSAGEM_INSTITUCIONAL,
            'TELEFONE': TELEFONE,
            'EMAIL': EMAIL,
            'WHATSAPP': WHATSAPP,
            'CIDADE': CIDADE,
            'BAIRRO': BAIRRO,
            'RUA': RUA,
            'NUMERO': NUMERO,
            'PIX': PIX,
            'FACEBOOK': FACEBOOK,
            'INSTAGRAM': INSTAGRAM,
            'YOUTUBE': YOUTUBE,
            'X': X,
            'TIKTOK': TIKTOK,

            'HORARIO_FUNCIONAMENTO_INICIO': HORARIO_FUNC_INICIO.strftime('%H:%M'),
            'HORARIO_FUNCIONAMENTO_FIM': HORARIO_FUNC_FIM.strftime('%H:%M'),
            'HORARIO_ALMOCO_INICIO': HORARIO_ALMOCO_INICIO.strftime('%H:%M'),
            'HORARIO_ALMOCO_FIM': HORARIO_ALMOCO_FIM.strftime('%H:%M')
        }

        return jsonify({"mensagem": 'Configurações', "configuracao": configuracao}), 200
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITA PARAMETRIZAÇÃO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/configuracoes/editar', methods=['PUT'])
def editar_configuracao():
    data = request.get_json()

    try:
        cur = con.cursor()

        # Atualizar diretamente a linha com ID fixo (assumindo que só existe essa)
        sql = """
        UPDATE CONFIGURACOES SET
            RAZAO_SOCIAL = ?,
            NOME_FANTASIA = ?,
            MENSAGEM_INSTITUCIONAL = ?,
            TELEFONE = ?,
            EMAIL = ?,
            WHATSAPP = ?,
            CIDADE = ?,
            BAIRRO = ?,
            RUA = ?,
            NUMERO = ?,
            PIX = ?,
            FACEBOOK = ?,
            INSTAGRAM = ?,
            YOUTUBE = ?,
            X = ?,
            TIKTOK = ?,
            HORARIO_FUNCIONAMENTO_INICIO = ?,
            HORARIO_FUNCIONAMENTO_FIM = ?,
            HORARIO_ALMOCO_INICIO = ?,
            HORARIO_ALMOCO_FIM = ?
        WHERE ID_CONFIGURACAO = 1
        """

        cur.execute(sql, (
            data.get("RAZAO_SOCIAL"),
            data.get("NOME_FANTASIA"),
            data.get("MENSAGEM_INSTITUCIONAL"),
            data.get("TELEFONE"),
            data.get("EMAIL"),
            data.get("WHATSAPP"),
            data.get("CIDADE"),
            data.get("BAIRRO"),
            data.get("RUA"),
            data.get("NUMERO"),
            data.get("PIX"),
            data.get("FACEBOOK"),
            data.get("INSTAGRAM"),
            data.get("YOUTUBE"),
            data.get("X"),
            data.get("TIKTOK"),
            data.get("HORARIO_FUNCIONAMENTO_INICIO"),
            data.get("HORARIO_FUNCIONAMENTO_FIM"),
            data.get("HORARIO_ALMOCO_INICIO"),
            data.get("HORARIO_ALMOCO_FIM")
        ))

        con.commit()
        cur.close()
        return jsonify({"mensagem": "Configuração atualizada com sucesso."})

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ADICIONA PARAMETRIZAÇÃO (VALORES) =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/configuracoes_valores', methods=['POST'])
def adicionar_configuracao_valores():
    try:
        dados = request.get_json()
        if not dados:
            return jsonify({"erro": "Dados JSON ausentes ou inválidos."}), 400

        # Validação
        campos = ['MAX_LIVROS', 'DIAS_EMPRESTIMO', 'VALOR_BASE', 'VALOR_ACRESCIMO']
        for campo in campos:
            if dados.get(campo) is None:
                return jsonify({"erro": f"Campo obrigatório '{campo}' ausente."}), 400

        # Data atual
        data_adicionado = datetime.now()

        # Inserção
        cur = con.cursor()
        cur.execute("""
            INSERT INTO VALORES (
                MAX_LIVROS, DIAS_EMPRESTIMO, DATA_ADICIONADO, VALOR_BASE, VALOR_ACRESCIMO
            ) VALUES (?, ?, ?, ?, ?)
        """, (
            int(dados.get('MAX_LIVROS')),
            int(dados.get('DIAS_EMPRESTIMO')),
            data_adicionado,
            float(dados.get('VALOR_BASE')),
            float(dados.get('VALOR_ACRESCIMO'))
        ))
        con.commit()
        cur.close()

        return jsonify({"mensagem": "Configuração adicionada com sucesso!"}), 201

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= RECUPERA PARAMETRIZAÇÃO (VALORES) =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/configuracoes_valores', methods=['GET'])
def obter_configuracao_valores():
    try:
        cur = con.cursor()

        # Buscar a última linha da tabela VALORES
        cur.execute("""
            SELECT ID_VALORES, MAX_LIVROS, DIAS_EMPRESTIMO, DATA_ADICIONADO, VALOR_BASE, VALOR_ACRESCIMO
            FROM VALORES
            ORDER BY DATA_ADICIONADO DESC, ID_VALORES DESC
            ROWS 1
        """)
        linha_valores = cur.fetchone()
        print("linha_valores:", linha_valores)

        # Inserir valor padrão se a tabela estiver vazia
        if not linha_valores:
            print("Tabela VALORES está vazia, adicionando padrão.")
            data_atual = datetime.now()
            cur.execute("""
                INSERT INTO VALORES (
                    MAX_LIVROS, DIAS_EMPRESTIMO, DATA_ADICIONADO, VALOR_BASE, VALOR_ACRESCIMO
                ) VALUES (?, ?, ?, ?, ?)
            """, (3, 15, data_atual, 3.50, 0.50))
            con.commit()

            cur.execute("""
                SELECT ID_VALORES, MAX_LIVROS, DIAS_EMPRESTIMO, DATA_ADICIONADO, VALOR_BASE, VALOR_ACRESCIMO
                FROM VALORES
                ORDER BY DATA_ADICIONADO DESC, ID_VALORES DESC
                ROWS 1
            """)
            linha_valores = cur.fetchone()
            print("Valor padrão inserido.")

        cur.close()

        # Atualizar app.config (opcional, só se quiser usar em outro lugar)
        app.config['ID_VALORES'] = linha_valores[0]
        app.config['MAX_LIVROS'] = linha_valores[1]
        app.config['DIAS_EMPRESTIMO'] = linha_valores[2]
        app.config['DATA_ADICIONADO'] = linha_valores[3]
        app.config['VALOR_BASE'] = linha_valores[4]
        app.config['VALOR_ACRESCIMO'] = linha_valores[5]

        # Retornar como JSON
        configuracao = {
            "ID_VALORES": linha_valores[0],
            "MAX_LIVROS": linha_valores[1],
            "DIAS_EMPRESTIMO": linha_valores[2],
            "DATA_ADICIONADO": linha_valores[3].strftime('%Y-%m-%d %H:%M:%S'),
            "VALOR_BASE": linha_valores[4],
            "VALOR_ACRESCIMO": linha_valores[5]
        }

        print("Configurações e valores carregados com sucesso!")
        return jsonify({"mensagem": "Configurações", "configuracao": configuracao}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PESQUISAR USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/usuarios', methods=['GET'])
def usuarios():
    try:
        # Criar cursor para consulta
        cur = con.cursor()

        # Consulta SQL para obter os dados da tabela USUARIO
        cur.execute(
            "SELECT id_usuario, nome, email, telefone, endereco, senha, tipo, ativo, data_nascimento FROM USUARIOS")
        usuarios = cur.fetchall()

        # Inicializar a lista de dicionários
        usuarios_dic = []
        for usuario in usuarios:
            id_usuario = usuario[0]

            caminho_imagem = f"Uploads/Usuarios/{id_usuario}.jpeg"

            if os.path.exists(caminho_imagem):
                print(f"A imagem {caminho_imagem} existe.")
            else:
                print(f"A imagem {caminho_imagem} não foi encontrada.")

                # Defina o caminho para a imagem padrão aqui, se necessário
                caminho_imagem = "foto-vazia.jpg"
                print(f"Usando imagem padrão: {caminho_imagem}")

            # Construir URL da imagem assumindo que será "id_livro.jpeg"
            imagem_url = url_for('static', filename=caminho_imagem, _external=True)

            usuarios_dic.append({
                'id_usuario': usuario[0],
                'nome': usuario[1],
                'email': usuario[2],
                'telefone': usuario[3],
                'endereco': usuario[4],
                'senha': usuario[5],
                'tipo': usuario[6],
                'ativo': usuario[7],
                'data_nascimento': usuario[8],
                'imagens': [imagem_url]
            })

        # Retornar os resultados como JSON
        return jsonify(mensagem='Lista de Usuários', usuarios=usuarios_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PESQUISAR USUARIO PELO ID =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/usuarios/<int:id>', methods=['GET'])
def usuario_por_id(id):
    try:
        # Criar cursor para consulta
        cur = con.cursor()

        # Consulta SQL para obter os dados do usuário com o ID especificado
        cur.execute(
            "SELECT id_usuario, nome, email, telefone, endereco, senha, tipo, ativo, data_nascimento FROM USUARIOS WHERE id_usuario = ?",
            (id,)
        )
        usuario = cur.fetchone()

        # Verificar se encontrou o usuário
        if not usuario:
            return jsonify(mensagem=f"Usuário com ID {id} não encontrado."), 404

        # Verificar se existe imagem correspondente
        caminho_imagem = f"Uploads/Usuarios/{id}.jpeg"
        if not os.path.exists(caminho_imagem):
            caminho_imagem = "foto-vazia.jpg"

        imagem_url = url_for('static', filename=caminho_imagem, _external=True)

        # Montar dicionário do usuário
        usuario_dic = {
            'id_usuario': usuario[0],
            'nome': usuario[1],
            'email': usuario[2],
            'telefone': usuario[3],
            'endereco': usuario[4],
            'senha': usuario[5],
            'tipo': usuario[6],
            'ativo': usuario[7],
            'data_nascimento': usuario[8],
            'imagens': [imagem_url]
        }

        return jsonify(mensagem='Usuário encontrado', usuario=usuario_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= VALIDAR A SENHA =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
def validar_senha(senha):
    if senha is None:
        return False, "A senha não pode ser None."

    if len(senha) < 8:
        return False, "A senha deve ter pelo menos 8 caracteres."

    tem_maiuscula = False
    tem_minuscula = False
    tem_numero = False
    tem_caracter_especial = False
    caracteres_especiais = "!@#$%^&*(),.?\":{}|<>"

    for char in senha:
        if char.isupper():
            tem_maiuscula = True
        elif char.islower():
            tem_minuscula = True
        elif char.isdigit():
            tem_numero = True
        elif char in caracteres_especiais:
            tem_caracter_especial = True

    if not tem_maiuscula:
        return False, "A senha deve conter pelo menos uma letra maiúscula."
    if not tem_minuscula:
        return False, "A senha deve conter pelo menos uma letra minúscula."
    if not tem_numero:
        return False, "A senha deve conter pelo menos um número."
    if not tem_caracter_especial:
        return False, "A senha deve conter pelo menos um caractere especial."

    return True, "Senha válida!"


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ADICIONAR USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/usuarios', methods=['POST'])
def criar_usuario():
    erros = []

    nome = request.form.get('nome')
    email = request.form.get('email')
    if email:
        email = email.lower().strip()
    telefone = request.form.get('telefone')
    endereco = request.form.get('endereco')
    senha = request.form.get('senha')
    tipo = request.form.get('tipo')
    ativo = request.form.get('ativo', True)
    data_nascimento = request.form.get('data_nascimento')
    imagem = request.files.get('imagem')  # Isso ainda está correto para a imagem, caso seja enviada de outra forma

    senha_valida, mensagem_senha = validar_senha(senha)
    if not senha_valida:
        erros.append({"mensagem": mensagem_senha, "campo": "senha"})

    senha = generate_password_hash(senha).decode('utf-8')

    cursor = con.cursor()


    # Verificar se o usuário já existe pelo e-mail
    cursor.execute("SELECT 1 FROM USUARIOS WHERE email = ?", (email,))
    if cursor.fetchone():
        erros.append({"mensagem": "Este e-mail já está cadastrado", "campo": "email"})

    # Verificar se o usuário já existe pelo telefone
    cursor.execute("SELECT 1 FROM USUARIOS WHERE telefone = ?", (telefone,))
    if cursor.fetchone():
        erros.append({"mensagem": "Este telefone já está cadastrado", "campo": "telefone"})

    if erros:
        cursor.close()
        # Retorna um JSON com uma chave "erros" que contém a lista
        return jsonify({"erros": erros}), 400

    # Inserir o novo usuário
    cursor.execute(""" 
        INSERT INTO USUARIOS (nome, email, telefone, endereco, senha, tipo, ativo, data_nascimento) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?) RETURNING ID_usuario
    """, (nome, email, telefone, endereco, senha, tipo, ativo, data_nascimento))

    id_usuario = cursor.fetchone()[0]
    con.commit()
    cursor.close()

    imagem_path = None
    nome_imagem = f"{id_usuario}.jpeg"
    pasta_destino = os.path.join(app.config['UPLOAD_FOLDER'], "Usuarios")
    os.makedirs(pasta_destino, exist_ok=True)
    imagem_path = os.path.join(pasta_destino, nome_imagem)

    if imagem:
        imagem.save(imagem_path)
    else:
        imagem_padrao = os.path.join(app.config['UPLOAD_FOLDER'], "foto-vazia.jpg")
        shutil.copy(imagem_padrao, imagem_path)

    return jsonify({
        'titulo': "Deu certo!",
        'texto': "Usuário cadastrado com sucesso",
        'message': "Usuário cadastrado com sucesso!",
        'usuario': {
            'nome': nome,
            'email': email,
            'telefone': telefone,
            'endereco': endereco,
            'senha': senha,
            'tipo': tipo,
            'ativo': ativo,
            'data_nascimento': data_nascimento,
            'imagem_path': imagem_path
        }
    }), 201


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITAR USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/usuarios/<int:id>', methods=['PUT'])
def editar_usuario(id):
    cursor = con.cursor()

    # Verificar se o usuário existe
    cursor.execute(
        "SELECT id_usuario, senha, email, nome, telefone, ativo, data_nascimento FROM USUARIOS WHERE id_usuario = ?",
        (id,))
    usuario_data = cursor.fetchone()

    senha_armazenada = usuario_data[1]
    email_armazenado = usuario_data[2]
    nome_armazenado = usuario_data[3]
    telefone_armazenado = usuario_data[4]
    ativo_armazenado = usuario_data[5]
    data_nascimento_armazenada = usuario_data[6]

    if not usuario_data:
        cursor.close()
        return jsonify({"erro": "Usuário não encontrado"}), 404

    # Caso o usuário exista, faça a atualização

    nome = request.form.get('nome')
    email = request.form.get('email')
    telefone = request.form.get('telefone')
    endereco = request.form.get('endereco')
    senha = request.form.get('senha')  # Lembre-se de criptografar antes de salvar
    tipo = request.form.get('tipo')
    ativo = request.form.get('ativo')
    data_nascimento = request.form.get('data_nascimento')
    imagem = request.files.get('imagem')

    if senha:
        senha_valida, mensagem_senha = validar_senha(senha)
        if not senha_valida:
            return jsonify({"erro": mensagem_senha}), 400
        senha = generate_password_hash(senha).decode('utf-8')
    else:
        senha = senha_armazenada

    if not email:
        email = email_armazenado

    if not data_nascimento:
        data_nascimento = data_nascimento_armazenada

    if not nome:
        nome = nome_armazenado

    if not telefone:
        telefone = telefone_armazenado

    if not ativo:
        ativo = ativo_armazenado
    # verifica o email
    cursor.execute("SELECT id_usuario FROM USUARIOS WHERE email = ? and id_usuario != ? ", (email, id))
    usuario_data = cursor.fetchone()
    if usuario_data:
        cursor.close()
        return jsonify({"erro": "Email já esta sendo usado ."}), 404

    # verifica o numero de telefone
    cursor.execute("SELECT id_usuario FROM USUARIOS WHERE telefone = ? and id_usuario != ? ", (telefone, id))
    usuario_data = cursor.fetchone()

    if usuario_data:
        cursor.close()
        return jsonify({"erro": "telefone já esta sendo usado ."}), 404

    cursor.execute("""
        UPDATE USUARIOS 
        SET nome = ?, email = ?, telefone = ?, endereco = ?, senha = ?, tipo = ?, ativo = ?, data_nascimento = ? 
        WHERE id_usuario = ?
    """, (nome, email, telefone, endereco, senha, tipo, ativo, data_nascimento, id))

    con.commit()
    imagem_path = None
    if imagem:
        nome_imagem = f"{id}.jpeg"
        pasta_destino = os.path.join(app.config['UPLOAD_FOLDER'], "Usuarios")
        os.makedirs(pasta_destino, exist_ok=True)
        imagem_path = os.path.join(pasta_destino, nome_imagem)
        imagem.save(imagem_path)

    cursor.close()

    return jsonify({
        'message': "Usuário atualizado com sucesso!",
        'usuario': {
            'id_usuario': id,
            'nome': nome,
            'email': email,
            'telefone': telefone,
            'endereco': endereco,
            'tipo': tipo,
            'ativo': ativo,
            'data_nascimento': data_nascimento,
            'imagem_path': imagem_path
        }
    })


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= DELETAR USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/usuarios/<int:id>', methods=['DELETE'])
def deletar_usuario(id):
    cursor = con.cursor()

    # Verificar se o usuário existe na tabela USUARIO
    cursor.execute("SELECT 1 FROM USUARIOS WHERE id_usuario = ?", (id,))
    if not cursor.fetchone():
        cursor.close()
        return jsonify({"erro": "Usuário não encontrado"}), 404

    # Excluir o usuário
    cursor.execute("DELETE FROM USUARIOS WHERE id_usuario = ?", (id,))
    con.commit()
    cursor.close()

    return jsonify({
        'message': "Usuário excluído com sucesso!",
        'id_usuario': id
    })


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PDF USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/usuarios/relatorio', methods=['GET'])
def gerar_relatorio_usuarios():
    filtro = request.args.get('filtro', 'todos')

    cursor = con.cursor()
    cursor.execute("""
        SELECT id_usuario, nome, email, telefone, endereco, tipo, ativo, data_nascimento
        FROM USUARIOS
        ORDER BY nome
    """)
    usuarios = cursor.fetchall()

    # Buscar IDs e valores de multas
    cursor.execute("SELECT id_usuario, valor_total FROM MULTAS WHERE status = 1")
    multas_data = cursor.fetchall()
    multas_ids = {id_: valor for id_, valor in multas_data}

    tipos_usuarios = {1: "Leitor", 2: "Bibliotecário", 3: "ADM"}

    usuarios_filtrados = []
    for u in usuarios:
        id_usuario, nome, email, telefone, endereco, tipo, ativo, data_nascimento = u
        is_multado = id_usuario in multas_ids

        if filtro == "todos":
            usuarios_filtrados.append(u)
        elif filtro == "adms" and tipo == 3:
            usuarios_filtrados.append(u)
        elif filtro == "bibliotecarios" and tipo == 2:
            usuarios_filtrados.append(u)
        elif filtro == "leitores" and tipo == 1:
            usuarios_filtrados.append(u)
        elif filtro == "inativos" and not ativo:
            usuarios_filtrados.append(u)
        elif filtro == "multados" and is_multado:
            usuarios_filtrados.append(u)

    mostrar_tipo = filtro not in ['adms', 'bibliotecarios', 'leitores']
    mostrar_status = filtro != 'inativos'
    mostrar_valor_multa = filtro == 'multados'
    mostrar_coluna_multado = filtro not in ['multados']

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 14)
    titulo = {
        'todos': 'Todos os Usuários',
        'adms': 'Administradores',
        'bibliotecarios': 'Bibliotecários',
        'leitores': 'Leitores',
        'inativos': 'Usuários Inativos',
        'multados': 'Usuários Multados'
    }.get(filtro, 'Usuários')
    pdf.cell(0, 10, f"Relatório - {titulo}", ln=True, align='C')
    pdf.ln(5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    # Cabeçalhos
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(40, 10, "Nome", 1)
    pdf.cell(50, 10, "E-mail", 1)
    if mostrar_tipo:
        pdf.cell(25, 10, "Tipo", 1)
    if mostrar_status:
        pdf.cell(20, 10, "Status", 1)
    pdf.cell(35, 10, "Data Nasc.", 1)
    if mostrar_coluna_multado:
        pdf.cell(20, 10, "Multado", 1)
    if mostrar_valor_multa:
        pdf.cell(20, 10, "Valor", 1)
    pdf.ln()

    pdf.set_font('Arial', '', 10)
    for u in usuarios_filtrados:
        id_usuario, nome, email, telefone, endereco, tipo, ativo, data_nascimento = u
        tipo_str = tipos_usuarios.get(tipo, "Desconhecido")
        ativo_str = "Ativo" if ativo else "Inativo"
        data_str = data_nascimento.strftime("%d/%m/%Y") if data_nascimento else "---"

        pdf.cell(40, 10, nome[:30], 1)
        pdf.cell(50, 10, email[:30], 1)
        if mostrar_tipo:
            pdf.cell(25, 10, tipo_str, 1)
        if mostrar_status:
            pdf.cell(20, 10, ativo_str, 1)
        pdf.cell(35, 10, data_str, 1)
        if mostrar_coluna_multado:
            multado_str = "Sim" if id_usuario in multas_ids else "Não"
            pdf.cell(20, 10, multado_str, 1)
        if mostrar_valor_multa:
            valor = multas_ids.get(id_usuario, 0)
            pdf.cell(20, 10, f"R${valor:.2f}", 1)

        pdf.ln()

    pdf.ln(5)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"Total de usuários no relatório: {len(usuarios_filtrados)}", ln=True, align='C')

    pdf_path = "relatorio_usuarios.pdf"
    pdf.output(pdf_path)
    cursor.close()
    return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= LOGIN =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
tentativas = 0

TITULOS_BEM_VINDO = [
    "Olá, {nome}!",
    "Bom te ver de novo, {nome}!",
    "Bom que você voltou, {nome}!",
    "Bem vindo de volta, {nome}!",
    "Quanto tempo, {nome}!"
]

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    senha = data.get('senha')
    global tentativas

    cursor = con.cursor()
    cursor.execute(
        'SELECT senha, id_usuario, nome, email, telefone, endereco, tipo, ativo, data_nascimento  FROM USUARIOS WHERE EMAIL = ?',
        (email,))
    resultado = cursor.fetchone()
    cursor.close()

    if not resultado:
        return jsonify({"erro": "Usuário não encontrado"}), 404

    senha_hash = resultado[0]
    id_usuario = resultado[1]
    nome = resultado[2]
    email = resultado[3]
    telefone = resultado[4]
    endereco = resultado[5]
    tipo = resultado[6]
    ativo = resultado[7]
    data_nascimento = resultado[8]

    if ativo == True:
        if check_password_hash(senha_hash, senha):
            token = generate_token(id_usuario)
            atualizar_status_agendamento()


            primeiro_nome = nome.split(' ')[0]
            titulo_template = random.choice(TITULOS_BEM_VINDO)
            titulo_personalizado = titulo_template.format(nome=primeiro_nome)
            mensagem_secundaria = "Login efetuado com sucesso!"

            return jsonify({
                "titulo": titulo_personalizado,
                "texto": mensagem_secundaria,
                'mensagem': 'Login efetuado.',
                'token': token,
                'id_usuario': id_usuario,
                'nome': nome,
                'email': email,
                'telefone': telefone,
                'endereco': endereco,
                'tipo': tipo,
                'ativo': ativo,
                'data_nascimento': data_nascimento
            }), 200


        else:
            tentativas += 1
            if tentativas == 3:
                cursor = con.cursor()
                cursor.execute("UPDATE USUARIOS SET ATIVO = 'false' WHERE id_usuario = ?", (id_usuario,))
                con.commit()
                cursor.close()

                corpo_email = f"""
                Olá <strong>{nome}</strong>,<br><br>
                Informamos que sua conta foi <strong>inativada</strong> devido a três (3) tentativas de login malsucedidas.<br>
                Por questões de segurança, sua conta foi inativada. Para reativá-la, entre em contato com o suporte.<br><br>
                <h3>Suporte</h3>
                Telefone: (xx) xxxx-xxxx<br>
                Ou envie um e-mail para <a href="mailto:support@bibliotecalibris.com">bibliotecalibris@gmail.com</a>.<br><br>
                """

                enviar_email(destinatario=email, assunto="🚨 Conta Conta por Tentativas Incorretas",
                             corpo_html=corpo_email, titulo="Conta Inativada")
                tentativas = 0

            return jsonify({"erro": "Senha incorreta"}), 401


    else:
        return jsonify({"erro": "Usuario inativo"}), 401


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= LOGOUT =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/logout', methods=['POST'])
def logout():
    token = request.headers.get('Autorization')

    if not token:
        return jsonify({"erro": "Token de autorização necessaria"}), 401

    # remove o bearer
    token = remover_bearer(token)

    try:
        # validar a assinatura e verificar a validade
        payload = jwt.decode(token, app.config['SECRET_KEY'], algoritms=['HS256'])

        # removendo o token do cliente
    except jwt.ExpiredSignatureError:
        return jsonify({"erro": "Token expirado"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"erro": "Token invalido"}), 401

    return jsonify({"message": "Logout bem-sucedido"}), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= MOSTRAR LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/acervo', methods=['GET'])
def acervo():
    print("ELE FEZ O SELECT ERRADO")
    try:
        # Criar cursor para consulta
        cur = con.cursor()

        # Consulta SQL com as novas colunas incluídas
        cur.execute("""
            SELECT 
                A.id_livro,
                A.titulo,
                A.autor,
                A.categoria,
                A.isbn,
                A.qtd_disponivel,
                A.sinopse,
                A.idioma,
                A.ano_publicado,
                A.qtd_total,
                A.avaliacao,
                A.editora,
                A.paginas,
                A.maior,
                A.local_fisico,
                (SELECT LIST(lt2.id_tag)
                 FROM LIVRO_TAGS lt2
                 WHERE lt2.id_livro = A.id_livro) AS tags_ids
            FROM ACERVO A
        """)

        acervo = cur.fetchall()

        # Inicializar a lista de dicionários
        acervo_dic = []

        for livro in acervo:
            id_livro = livro[0]
            tags_ids = livro[15]  # Posição do campo LIST()

            imagem_url = url_for('static', filename=f"Uploads/Livros/{id_livro}.jpeg", _external=True)

            # Adicionar os dados do livro ao dicionário
            acervo_dic.append({
                'id_livro': id_livro,
                'titulo': livro[1],
                'autor': livro[2],
                'categoria': livro[3],
                'isbn': livro[4],
                'qtd_disponivel': livro[5],
                'sinopse': livro[6],
                'idioma': livro[7],
                'ano_publicado': livro[8],
                'qtd_total': livro[9],
                'avaliacao': livro[10],
                'editora': livro[11],
                'paginas': livro[12],
                'maior': livro[13],
                'local_fisico': livro[14],
                'tags_ids': tags_ids,
                'imagens': [imagem_url]
            })

        return jsonify(mensagem='Lista do Acervo', acervo=acervo_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados: {e}"), 500

@app.route('/acervo/do-mesmo-autor/<int:id_livro_inicial>', methods=['GET'])
def acervo_do_mesmo_autor(id_livro_inicial):
    """
    Recebe o ID de um livro e retorna todos os outros livros do mesmo autor.
    """
    try:
        cur = con.cursor()

        # --- A MÁGICA ESTÁ AQUI, NA CONSULTA SQL ---
        # Usamos uma subquery para encontrar o nome do autor baseado no ID inicial
        # e depois usamos esse nome para buscar todos os outros livros.
        sql_query = """
            SELECT 
                A.id_livro, A.titulo, A.autor, A.categoria, A.isbn,
                A.qtd_disponivel, A.sinopse, A.idioma, A.ano_publicado,
                A.qtd_total, A.avaliacao, A.editora, A.paginas,
                A.maior, A.local_fisico,
                (SELECT LIST(lt2.id_tag)
                 FROM LIVRO_TAGS lt2
                 WHERE lt2.id_livro = A.id_livro) AS tags_ids
            FROM ACERVO A
            WHERE 
                -- 1. Encontra livros cujo autor é igual ao autor do livro com o ID inicial
                A.autor = (SELECT A2.autor FROM ACERVO A2 WHERE A2.id_livro = ?)
                -- 2. E exclui o próprio livro inicial da lista de resultados
                AND A.id_livro != ?
        """

        # Executamos a consulta passando o ID inicial duas vezes:
        # uma para a subquery e outra para a condição de exclusão.
        cur.execute(sql_query, (id_livro_inicial, id_livro_inicial))

        livros_do_autor = cur.fetchall()

        # Mesmo que não encontre outros livros, não é um erro 404.
        # Uma lista vazia é uma resposta válida.
        # A verificação de 404 pode ser removida ou mantida, conforme preferir.
        if not livros_do_autor:
            return jsonify(mensagem=f"Nenhum outro livro encontrado do mesmo autor do livro ID: {id_livro_inicial}",
                           acervo=[]), 200

        # O código para montar o JSON de resposta é exatamente o mesmo,
        # garantindo a consistência da sua API.
        acervo_dic = []
        for livro in livros_do_autor:
            id_livro_loop = livro[0]
            tags_ids = livro[15]
            imagem_url = url_for('static', filename=f"Uploads/Livros/{id_livro_loop}.jpeg", _external=True)
            acervo_dic.append({
                'id_livro': id_livro_loop,
                'titulo': livro[1],
                'autor': livro[2],
                'categoria': livro[3],
                'isbn': livro[4],
                'qtd_disponivel': livro[5],
                'sinopse': livro[6],
                'idioma': livro[7],
                'ano_publicado': livro[8],
                'qtd_total': livro[9],
                'avaliacao': livro[10],
                'editora': livro[11],
                'paginas': livro[12],
                'maior': livro[13],
                'local_fisico': livro[14],
                'tags_ids': tags_ids,
                'imagens': [imagem_url]
            })

        return jsonify(mensagem=f"Outros livros do mesmo autor", acervo=acervo_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao buscar livros do mesmo autor: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= MOSTRAR LIVRO PELO ID =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/acervo/<int:id_livro>', methods=['GET'])
def acervo_por_id(id_livro):
    try:
        cur = con.cursor()
        cur.execute("""
            SELECT 
                A.id_livro,
                A.titulo,
                A.autor,
                A.categoria,
                A.isbn,
                A.qtd_disponivel,
                A.sinopse,
                A.idioma,
                A.ano_publicado,
                A.qtd_total,
                A.avaliacao,
                A.editora,
                A.paginas,
                A.maior,
                A.local_fisico,
                (SELECT LIST(lt2.id_tag)
                 FROM LIVRO_TAGS lt2
                 WHERE lt2.id_livro = A.id_livro) AS tags_ids
            FROM ACERVO A
            WHERE A.id_livro = ?
        """, (id_livro,))

        livro = cur.fetchone()

        if not livro:
            return jsonify(mensagem="Livro não encontrado"), 404

        imagem_url = url_for('static', filename=f"Uploads/Livros/{livro[0]}.jpeg", _external=True)

        livro_dic = {
            'id_livro': livro[0],
            'titulo': livro[1],
            'autor': livro[2],
            'categoria': livro[3],
            'isbn': livro[4],
            'qtd_disponivel': livro[5],
            'sinopse': livro[6],
            'idioma': livro[7],
            'ano_publicado': livro[8],
            'qtd_total': livro[9],
            'avaliacao': livro[10],
            'editora': livro[11],
            'paginas': livro[12],
            'maior': livro[13],
            'local_fisico': livro[14],
            'tags_ids': livro[15],
            'imagens': [imagem_url]
        }

        return jsonify(mensagem="Livro encontrado", livro=livro_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PESQUISAR LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
@app.route('/acervo/filtro', methods=['GET'])
def acervo_pesquisa():
    try:
        cur = con.cursor()

        # Parâmetros da URL
        busca = request.args.get('busca', '').strip().lower()
        genero = request.args.get('genero', type=int)
        disponibilidade = request.args.get('disponibilidade', '').strip().lower()
        min_paginas = request.args.get('minPaginas', type=int, default=0)
        min_votos = request.args.get('minVotos', type=int, default=0)
        palavra_chave = request.args.get('palavraChave', '').strip().lower()

        # Query base
        base_query = """
            SELECT A.id_livro, A.titulo, A.autor, A.categoria, A.isbn, A.qtd_disponivel,
                   A.sinopse, A.idioma, A.ano_publicado, A.qtd_total, A.avaliacao, 
                   A.palavra_chave, A.paginas
            FROM ACERVO A
            WHERE 1=1
        """
        params = []

        # 🔍 Filtro de busca (título ou autor)
        if busca:
            base_query += " AND (LOWER(A.titulo) LIKE ? OR LOWER(A.autor) LIKE ?)"
            params.extend([f"%{busca}%", f"%{busca}%"])

        # 🔑 Filtro de palavra-chave (título, autor, sinopse, tags)
        if palavra_chave:
            base_query += """
                AND (
                    LOWER(A.titulo) LIKE ? OR
                    LOWER(A.autor) LIKE ? OR
                    LOWER(A.sinopse) LIKE ? OR
                    EXISTS (
                        SELECT 1 
                        FROM LIVRO_TAGS LT 
                        JOIN TAGS T ON LT.id_tag = T.id_tag 
                        WHERE LT.id_livro = A.id_livro 
                        AND LOWER(T.NOME_TAG) LIKE ?
                    )
                )
            """
            termo = f"%{palavra_chave}%"
            params.extend([termo, termo, termo, termo])

        # 🎭 Filtro de gênero (via relação com TAGS)
        if genero:
            base_query += """
                AND EXISTS (
                    SELECT 1 FROM LIVRO_TAGS LT_gen 
                    WHERE LT_gen.id_livro = A.id_livro 
                    AND LT_gen.id_tag = ?
                )
            """
            params.append(genero)

        # 📦 Filtro por disponibilidade
        if disponibilidade == 'disponiveis':
            base_query += " AND A.qtd_disponivel > 0"
        elif disponibilidade == 'emprestados':
            base_query += " AND A.qtd_total > 0 AND A.qtd_disponivel = 0"

        # 📘 Filtro por páginas mínimas
        if min_paginas > 0:
            base_query += " AND A.paginas >= ?"
            params.append(min_paginas)

        # ⭐ Filtro por avaliação mínima
        if min_votos > 0:
            base_query += " AND A.avaliacao >= ?"
            params.append(min_votos)

        base_query += " ORDER BY A.titulo ASC"

        # Execução da query
        cur.execute(base_query, tuple(params))
        acervo_rows = cur.fetchall()
        cur.close()

        # Montagem do dicionário de retorno
        acervo_dic = []
        for livro in acervo_rows:
            id_livro = livro[0]

            # Busca das tags do livro
            tags_cur = con.cursor()
            tags_cur.execute("""
                SELECT T.id_tag, T.NOME_TAG
                FROM TAGS T
                JOIN LIVRO_TAGS LT ON T.id_tag = LT.id_tag
                WHERE LT.id_livro = ?
            """, (id_livro,))
            tags = tags_cur.fetchall()
            tags_cur.close()

            tags_formatadas = [{'id': tag[0], 'nome': tag[1]} for tag in tags]

            acervo_dic.append({
                'id_livro': livro[0],
                'titulo': livro[1],
                'autor': livro[2],
                'categoria': livro[3],
                'isbn': livro[4],
                'qtd_disponivel': livro[5] if livro[5] is not None else 0,
                'sinopse': livro[6],
                'idioma': livro[7],
                'ano_publicado': livro[8],
                'qtd_total': livro[9],
                'avaliacao': livro[10] if livro[10] is not None else 0,
                'palavra_chave': livro[11],
                'paginas': livro[12] if livro[12] is not None else 0,
                'tags': tags_formatadas
            })

        return jsonify({'acervo': acervo_dic, 'mensagem': 'Lista do Acervo'})

    except Exception as e:
        print(f"Erro no backend: {e}")
        return jsonify({'erro': str(e)}), 500

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ADICIONAR LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/acervo', methods=['POST'])
def criar_livro():
    token = request.headers.get('Authorization')
    if not token:
        return jsonify({'mensagem': 'Token de autenticação necessário'}), 401

    token = remover_bearer(token)

    try:
        payload = jwt.decode(token, senha_secreta, algorithms=['HS256'])
        id_usuario = payload['id_usuario']
    except jwt.ExpiredSignatureError:
        return jsonify({'mensagem': 'Token expirado'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'mensagem': 'Token inválido'}), 401

    if not verifica_adm(id_usuario):
        return jsonify({'erro': 'Página restrita ao bobao'}), 401

    # Dados do livro
    titulo = request.form.get('titulo')
    autor = request.form.get('autor')
    categoria = request.form.get('categoria')
    isbn = request.form.get('isbn')
    qtd_disponivel = request.form.get('qtd_disponivel')
    qtd_total = request.form.get('qtd_total')
    sinopse = request.form.get('sinopse')
    idioma = request.form.get('idioma')
    ano_publicado = request.form.get('ano_publicado')
    editora = request.form.get('editora')
    paginas = request.form.get('paginas')
    maior = request.form.get('maior_idade') == 'on'  # Pode vir como string, tipo "true"
    local_fisico = request.form.get('local_fisico')
    id_tag = request.form.getlist('id_tag[]')
    imagem = request.files.get('imagem')

    # Convertendo "maior" para boolean (se necessário)
    if isinstance(maior, str):
        maior = maior.lower() in ['1', 'true', 'sim']

    cursor = con.cursor()

    # Verificar por título
    cursor.execute("SELECT 1 FROM ACERVO WHERE titulo = ?", (titulo,))
    if cursor.fetchone():
        return jsonify({"erro": "Livro já cadastrado com esse título"}), 400

    # Verificar por ISBN
    cursor.execute("SELECT 1 FROM ACERVO WHERE isbn = ?", (isbn,))
    if cursor.fetchone():
        return jsonify({"erro": "Já existe um livro cadastrado com esse ISBN"}), 400

    # Inserção com as novas colunas
    cursor.execute("""
        INSERT INTO ACERVO (
            titulo, autor, categoria, isbn, qtd_total, sinopse, idioma, 
            ano_publicado, qtd_disponivel, editora, paginas, maior, local_fisico
        ) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) 
        RETURNING ID_livro
    """, (
        titulo, autor, categoria, isbn, qtd_total, sinopse, idioma,
        ano_publicado, qtd_disponivel, editora, paginas, maior, local_fisico
    ))

    id_livro = cursor.fetchone()[0]
    con.commit()

    # Salvar imagem
    imagem_path = None
    if imagem:
        nome_imagem = f"{id_livro}.jpeg"
        pasta_destino = os.path.join(app.config['UPLOAD_FOLDER'], "Livros")
        os.makedirs(pasta_destino, exist_ok=True)
        imagem_path = os.path.join(pasta_destino, nome_imagem)
        imagem.save(imagem_path)

    # Vínculo com tags
    cur = con.cursor()

    if len(id_tag) == 1 and ',' in id_tag[0]:
        id_tag = id_tag[0].split(',')

    id_tag = [tag.strip() for tag in id_tag]

    try:
        for tag in id_tag:
            if tag.isdigit():
                id_tag_inserida = int(tag)
            else:
                cur.execute("SELECT ID_TAG FROM TAGS WHERE nome_tag = ?", (tag,))
                existing_tag = cur.fetchone()

                if existing_tag:
                    id_tag_inserida = existing_tag[0]
                else:
                    cur.execute("INSERT INTO TAGS (nome_tag) VALUES (?)", (tag,))
                    cur.execute("SELECT FIRST 1 ID_TAG FROM TAGS ORDER BY ID_TAG DESC")
                    id_tag_inserida = cur.fetchone()[0]

            cur.execute("INSERT INTO LIVRO_TAGS (id_livro, id_tag) VALUES (?, ?)", (id_livro, id_tag_inserida))
        con.commit()

    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        con.rollback()
    finally:
        cur.close()

    return jsonify({
        'message': "Livro cadastrado com sucesso!",
        'livro': {
            'titulo': titulo,
            'autor': autor,
            'categoria': categoria,
            'isbn': isbn,
            'qtd_disponivel': qtd_disponivel,
            'qtd_total': qtd_total,
            'sinopse': sinopse,
            'idioma': idioma,
            'ano_publicado': ano_publicado,
            'editora': editora,
            'paginas': paginas,
            'maior': maior,
            'local_fisico': local_fisico,
            'tags': id_tag,
            'imagem_path': imagem_path
        }
    }), 201


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITAR LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/acervo/<int:id>', methods=['PUT'])
def editar_livro(id):
    try:
        cursor = con.cursor()

        # Verificar se o livro existe no acervo e obter qtd_total atual
        cursor.execute("SELECT qtd_total, qtd_disponivel FROM ACERVO WHERE id_livro = ?", (id,))
        livro_data = cursor.fetchone()

        if not livro_data:
            cursor.close()
            return jsonify({"erro": "Livro não encontrado"}), 404

        qtd_total_atual = livro_data[0]
        qtd_disponivel_atual = livro_data[1]

        # Captura dos dados da requisição
        titulo = request.form.get('titulo')
        autor = request.form.get('autor')
        categoria = request.form.get('categoria')
        isbn = request.form.get('isbn')
        qtd_total_novo = request.form.get('qtd_total')
        qtd_disponivel = request.form.get('qtd_disponivel')
        sinopse = request.form.get('sinopse')
        idioma = request.form.get('idioma')
        ano_publicado = request.form.get('ano_publicado')
        imagem = request.files.get('imagem')
        editora = request.form.get('editora')
        paginas = request.form.get('paginas')
        maior = request.form.get('maior_idade') == 'on'
        local_fisico = request.form.get('local_fisico')

        # Conversão de tipos
        qtd_total_novo = int(qtd_total_novo) if qtd_total_novo else qtd_total_atual
        qtd_disponivel = int(qtd_disponivel) if qtd_disponivel else qtd_disponivel_atual
        if isinstance(maior, str):
            maior = maior.lower() in ['1', 'true', 'sim']

        # Verificar duplicidade de título (em outro livro)
        cursor.execute("SELECT 1 FROM ACERVO WHERE titulo = ? AND id_livro != ?", (titulo, id))
        if cursor.fetchone():
            return jsonify({"erro": "Já existe outro livro com esse título"}), 400

        # Verificar duplicidade de ISBN (em outro livro)
        cursor.execute("SELECT 1 FROM ACERVO WHERE isbn = ? AND id_livro != ?", (isbn, id))
        if cursor.fetchone():
            return jsonify({"erro": "Já existe outro livro com esse ISBN"}), 400

        # Ajustar qtd_disponivel se qtd_total mudou
        if qtd_total_novo != qtd_total_atual:
            diferenca = qtd_total_novo - qtd_total_atual
            qtd_disponivel += diferenca
            if qtd_disponivel < 0:
                qtd_disponivel = 0

        # Atualização no banco com as novas colunas
        cursor.execute("""
            UPDATE ACERVO 
            SET titulo = ?, autor = ?, categoria = ?, isbn = ?, qtd_total = ?, 
                qtd_disponivel = ?, sinopse = ?, idioma = ?, ano_publicado = ?, 
                editora = ?, paginas = ?, maior = ?, local_fisico = ?
            WHERE id_livro = ?
        """, (
            titulo, autor, categoria, isbn, qtd_total_novo,
            qtd_disponivel, sinopse, idioma, ano_publicado,
            editora, paginas, maior, local_fisico, id
        ))

        con.commit()

        # Atualizar imagem se enviada
        imagem_path = None
        if imagem:
            nome_imagem = f"{id}.jpeg"
            pasta_destino = os.path.join(app.config['UPLOAD_FOLDER'], "Livros")
            os.makedirs(pasta_destino, exist_ok=True)
            imagem_path = os.path.join(pasta_destino, nome_imagem)
            imagem.save(imagem_path)

        # ---------------- ATUALIZAR TAGS ASSOCIADAS ----------------
        raw_tags = request.form.get('tags_ids', '')
        id_tags = [tag.strip() for tag in raw_tags.split(',') if tag.strip()]

        cursor.execute("DELETE FROM LIVRO_TAGS WHERE id_livro = ?", (id,))

        for tag in id_tags:
            if tag.isdigit():
                id_tag_inserida = int(tag)
            else:
                cursor.execute("SELECT ID_TAG FROM TAGS WHERE nome_tag = ?", (tag,))
                existing_tag = cursor.fetchone()
                if existing_tag:
                    id_tag_inserida = existing_tag[0]
                else:
                    cursor.execute("INSERT INTO TAGS (nome_tag) VALUES (?)", (tag,))
                    cursor.execute("SELECT FIRST 1 ID_TAG FROM TAGS ORDER BY ID_TAG DESC")
                    id_tag_inserida = cursor.fetchone()[0]

            cursor.execute("INSERT INTO LIVRO_TAGS (id_livro, id_tag) VALUES (?, ?)", (id, id_tag_inserida))

        con.commit()
        cursor.close()

        return jsonify({
            'message': "Livro atualizado com sucesso!",
            'livro': {
                'id_livro': id,
                'titulo': titulo,
                'autor': autor,
                'categoria': categoria,
                'isbn': isbn,
                'qtd_total': qtd_total_novo,
                'qtd_disponivel': qtd_disponivel,
                'sinopse': sinopse,
                'idioma': idioma,
                'ano_publicado': ano_publicado,
                'editora': editora,
                'paginas': paginas,
                'maior': maior,
                'local_fisico': local_fisico,
                'imagem_path': imagem_path,
                'tags': id_tags
            }
        })

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= DELETAR LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/acervo/<int:id>', methods=['DELETE'])
def deletar_livro(id):
    cursor = con.cursor()

    # Verifica se o livro existe
    cursor.execute("SELECT 1 FROM ACERVO WHERE id_livro = ?", (id,))
    if not cursor.fetchone():
        cursor.close()
        return jsonify({"erro": "Livro não encontrado"}), 404

    # Verifica se o livro já teve algum empréstimo
    cursor.execute("SELECT 1 FROM RESERVAS WHERE id_livro = ?", (id,))
    if cursor.fetchone():
        cursor.close()
        return jsonify({"erro": "Este livro já teve um empréstimo e não pode ser excluído."}), 400

    # Exclui o livro
    cursor.execute("DELETE FROM ACERVO WHERE id_livro = ?", (id,))
    con.commit()
    cursor.close()

    return jsonify({
        'message': "Livro excluído com sucesso!",
        'id_livro': id
    })


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PDF LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
# --- FUNÇÃO DE FILTRAGEM REUTILIZÁVEL - VERSÃO FINAL ---
def construir_clausulas_filtro(args):
    """
    Lê os argumentos da requisição e constrói a cláusula WHERE da SQL e os parâmetros.
    Retorna uma tupla: (string_where, lista_de_parametros)
    """
    conditions = []
    params = []

    # Mapeamento para filtros simples (LIKE)
    # ATENÇÃO: 'avaliacao' foi REMOVIDO daqui para ser tratado como um caso especial.
    mapeamento_filtros = {
        'autor': 'A.autor LIKE ?',
        'categoria': 'A.categoria LIKE ?',
        'editora': 'A.editora LIKE ?',
    }

    # Processa filtros simples
    for nome_filtro, condicao_sql in mapeamento_filtros.items():
        valor = args.get(nome_filtro)
        if valor:
            conditions.append(condicao_sql)
            params.append(f"%{valor}%")

    # --- TRATAMENTO ESPECIAL PARA AVALIAÇÃO (A CORREÇÃO PRINCIPAL) ---
    valor_avaliacao = args.get('avaliacao')
    if valor_avaliacao:
        try:
            # Converte o valor para número
            rating = float(valor_avaliacao)

            # Se a avaliação for 5, busca por uma correspondência exata.
            if rating == 5.0:
                conditions.append("A.avaliacao = ?")
                params.append(rating)
            # Para qualquer outra avaliação (1, 2, 3, 4), busca por um intervalo.
            # Ex: Se rating for 3, busca por >= 3 E < 4.
            else:
                conditions.append("A.avaliacao >= ? AND A.avaliacao < ?")
                params.extend([rating, rating + 1])
        except (ValueError):
            # Ignora o filtro se o valor não for um número válido
            pass

    # Processa filtro de gênero (EXISTS)
    if args.get('genero'):
        conditions.append("EXISTS (SELECT 1 FROM LIVRO_TAGS lt WHERE lt.id_livro = A.id_livro AND lt.id_tag = ?)")
        params.append(args.get('genero'))

    # Processa filtro de páginas (BETWEEN / >=)
    if args.get('paginas'):
        try:
            min_max = args.get('paginas').split('-')
            if len(min_max) == 2 and min_max[0] and min_max[1]:
                conditions.append("A.paginas BETWEEN ? AND ?")
                params.extend([int(min_max[0]), int(min_max[1])])
            elif len(min_max) == 2 and min_max[0] and not min_max[1]:
                conditions.append("A.paginas >= ?")
                params.append(int(min_max[0]))
        except (ValueError, IndexError):
            pass

    # Monta a string WHERE final
    where_clause = ""
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    return where_clause, params


# --- ROTA DE BUSCA AO VIVO (AGORA MAIS SIMPLES) ---
@app.route('/acervo/filtro', methods=['GET'])
def filtro_acervo():
    try:
        cur = con.cursor()

        # SQL base para a busca ao vivo
        sql_base = """
            SELECT A.id_livro, A.titulo, A.autor, A.categoria, A.isbn, A.qtd_disponivel, 
                   A.sinopse, A.idioma, A.ano_publicado, A.qtd_total, A.avaliacao, 
                   A.editora, A.paginas, A.maior, A.local_fisico,
                   (SELECT LIST(lt2.id_tag) FROM LIVRO_TAGS lt2 WHERE lt2.id_livro = A.id_livro) AS tags_ids
            FROM ACERVO A
        """

        # Usa a função reutilizável para obter os filtros
        where_clause, params = construir_clausulas_filtro(request.args)

        # Combina tudo
        sql_final = f"{sql_base} {where_clause} ORDER BY A.titulo"

        cur.execute(sql_final, tuple(params))
        acervo_db = cur.fetchall()

        # O resto do código para montar o JSON continua o mesmo...
        acervo_dic = []
        for livro in acervo_db:
            id_livro = livro[0]
            imagem_url = url_for('static', filename=f"Uploads/Livros/{id_livro}.jpeg", _external=True)
            acervo_dic.append(
                {'id_livro': id_livro, 'titulo': livro[1], 'autor': livro[2], 'categoria': livro[3], 'isbn': livro[4],
                 'qtd_disponivel': livro[5], 'sinopse': livro[6], 'idioma': livro[7], 'ano_publicado': livro[8],
                 'qtd_total': livro[9], 'avaliacao': livro[10], 'editora': livro[11], 'paginas': livro[12],
                 'maior': livro[13], 'local_fisico': livro[14], 'tags_ids': livro[15], 'imagens': [imagem_url]})

        return jsonify(mensagem='Lista do Acervo', acervo=acervo_dic)

    except Exception as e:
        print(f"ERRO NA ROTA /acervo/filtro: {e}")
        return jsonify(mensagem=f"Erro ao consultar o banco de dados."), 500


# =========================================================================
#  NOVA ROTA OTIMIZADA PARA TOTAIS DO ACERVO
# =========================================================================
@app.route('/acervo/estatisticas/geral', methods=['GET'])
def get_estatisticas_gerais_acervo():
    try:
        cur = con.cursor()

        # 1. Conta o total de livros no acervo
        cur.execute("SELECT COUNT(id_livro) FROM ACERVO")
        total_livros = cur.fetchone()[0]

        # 2. Conta os livros emprestados (Status = 2)
        cur.execute("SELECT COUNT(ID_LIVRO) FROM RESERVAS WHERE STATUS = 2")
        total_emprestados = cur.fetchone()[0]

        cur.close()

        # Retorna ambos os valores em um único objeto JSON
        return jsonify({
            'total_livros': total_livros,
            'total_emprestados': total_emprestados
        })

    except Exception as e:
        print(f"ERRO NA ROTA /acervo/estatisticas/geral: {e}")
        return jsonify(mensagem="Erro ao buscar estatísticas gerais."), 500


def criar_pdf_livros(lista_de_livros, titulo_relatorio, sumario_info):
    """
    Função reutilizável que recebe uma lista de livros e informações de cabeçalho
    e retorna o caminho para o arquivo PDF gerado.
    """
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # --- CABEÇALHO DO PDF ---
    pdf.image('static/uploads/logo.png', x=10, y=8, w=30)
    pdf.set_font('Arial', 'B', 18)
    pdf.cell(0, 10, titulo_relatorio, ln=True, align='C')
    pdf.set_font('Arial', '', 8)
    data_hoje = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    pdf.cell(0, 5, f"Gerado em: {data_hoje}", ln=True, align='C')
    pdf.ln(15)

    # Bloco de informações (sumário ou total)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 7, sumario_info['titulo'], ln=True)
    pdf.set_font('Arial', '', 9)
    pdf.multi_cell(0, 5, sumario_info['conteudo'], border=1, align='L')
    pdf.ln(5)

    # --- TABELA DE DADOS ---
    colunas = [("Título", 70), ("Autor", 38), ("ISBN", 28), ("Idioma", 14), ("Ano", 18), ("Disp/Total", 19)]

    def desenhar_cabecalho_tabela():
        pdf.set_font('Arial', 'B', 9)
        pdf.set_fill_color(230, 230, 230)
        y_cabecalho = pdf.get_y()
        x_cabecalho = pdf.get_x()
        for titulo, largura in colunas:
            pdf.set_xy(x_cabecalho, y_cabecalho)
            pdf.multi_cell(w=largura, h=5, txt=titulo, border=1, align='C', fill=True)
            x_cabecalho += largura
        pdf.set_y(y_cabecalho + 10)

    desenhar_cabecalho_tabela()

    altura_celula = 5
    for livro_tupla in lista_de_livros:
        pdf.set_font('Arial', '', 8)
        (titulo, autor, idioma, ano, disp, total, isbn) = livro_tupla
        dados = [str(titulo), str(autor), str(isbn), str(idioma), str(ano), f"{disp or 0}/{total or 0}"]
        dados_enc = [d.encode('latin-1', 'ignore').decode('latin-1') for d in dados]

        altura_max_linha = altura_celula
        for i in range(len(colunas)):
            num_linhas = len(pdf.multi_cell(w=colunas[i][1], h=altura_celula, txt=dados_enc[i], split_only=True))
            altura_max_linha = max(altura_max_linha, num_linhas * altura_celula)

        if pdf.get_y() + altura_max_linha > pdf.page_break_trigger:
            pdf.add_page()
            desenhar_cabecalho_tabela()

        y_linha = pdf.get_y()
        x_pos = pdf.get_x()
        for i in range(len(colunas)):
            pdf.set_xy(x_pos, y_linha)
            pdf.multi_cell(w=colunas[i][1], h=altura_max_linha, txt=dados_enc[i], border=1, align='L')
            x_pos += colunas[i][1]
        pdf.set_y(y_linha + altura_max_linha)

    pdf_path = f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(pdf_path)
    return pdf_path


def criar_excel_livros(lista_de_livros, titulo_relatorio, sumario_info):
    """
    Recebe uma lista de livros e gera um arquivo Excel (.xlsx).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Livros"

    # --- ESTILOS ---
    font_titulo = Font(name='Calibri', size=18, bold=True, color='2F5496')
    font_subtitulo = Font(name='Calibri', size=9, italic=True)
    font_cabecalho_tabela = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_cabecalho_tabela = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal='center', vertical='center')

    # --- CABEÇALHO DO ARQUIVO ---
    ws.merge_cells('A1:G1')
    ws['A1'] = titulo_relatorio
    ws['A1'].font = font_titulo
    ws['A1'].alignment = alignment_center

    ws.merge_cells('A2:G2')
    data_hoje = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws['A2'] = f"Gerado em: {data_hoje}"
    ws['A2'].font = font_subtitulo
    ws['A2'].alignment = alignment_center

    # --- SUMÁRIO ---
    ws.merge_cells('A4:G4')
    ws['A4'] = sumario_info['titulo']
    ws['A4'].font = Font(name='Calibri', size=10, bold=True)

    ws.merge_cells('A5:G6')
    ws['A5'] = sumario_info['conteudo']
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')

    # --- TABELA DE DADOS ---
    linha_atual = 8

    # Cabeçalho da Tabela
    colunas = ["Título", "Autor", "ISBN", "Idioma", "Ano", "Disponível", "Total"]
    ws.append(colunas)
    for col in range(1, len(colunas) + 1):
        cell = ws.cell(row=linha_atual, column=col)
        cell.font = font_cabecalho_tabela
        cell.fill = fill_cabecalho_tabela
        cell.alignment = alignment_center
    linha_atual += 1

    # Dados da Tabela
    for livro_tupla in lista_de_livros:
        # A ordem das colunas na tupla deve corresponder à sua consulta SQL
        (titulo, autor, idioma, ano, disp, total, isbn) = livro_tupla
        # Monta a linha na ordem correta das colunas
        dados_linha = [titulo, autor, isbn, idioma, ano, disp or 0, total or 0]
        ws.append(dados_linha)

    # Ajuste automático da largura das colunas
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Salva o arquivo e retorna o caminho
    excel_path = f"relatorio_livros_excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(excel_path)
    return excel_path


# ROTA DE RELATÓRIO FILTRADO
@app.route('/livros/relatorio', methods=['GET'])
def gerar_relatorio_livros():
    try:
        cur = con.cursor()

        # Monta o sumário dos filtros (código que você já tinha)
        filtros_usados_desc = []
        nomes_amigaveis = {'autor': 'Autor', 'categoria': 'Categoria', 'editora': 'Editora', 'avaliacao': 'Avaliação',
                           'paginas': 'Páginas'}
        for chave, nome in nomes_amigaveis.items():
            valor = request.args.get(chave)
            if valor:
                if chave == 'paginas':
                    min_max = valor.split('-')
                    desc = ""
                    if len(min_max) == 2 and min_max[0] and min_max[1]:
                        desc = f"de {min_max[0]} a {min_max[1]}"
                    elif len(min_max) == 2 and min_max[0]:
                        desc = f"a partir de {min_max[0]}"
                    if desc: filtros_usados_desc.append(f"{nome}: {desc}")
                else:
                    filtros_usados_desc.append(f"{nome}: {valor}")
        genero_id = request.args.get('genero')
        if genero_id:
            cur.execute("SELECT nome_tag FROM TAGS WHERE id_tag = ?", (genero_id,))
            resultado_tag = cur.fetchone()
            if resultado_tag: filtros_usados_desc.append(f"Gênero: {resultado_tag[0]}")
        sumario_filtros = ", ".join(filtros_usados_desc) if filtros_usados_desc else "Nenhum filtro aplicado."

        # Busca os livros filtrados
        sql_base = "SELECT A.titulo, A.autor, A.idioma, A.ano_publicado, A.qtd_disponivel, A.qtd_total, A.isbn FROM ACERVO A"
        where_clause, params = construir_clausulas_filtro(request.args)
        sql_final = f"{sql_base} {where_clause} ORDER BY A.titulo"
        cur.execute(sql_final, tuple(params))
        livros_filtrados = cur.fetchall()
        cur.close()

        # Prepara as informações para a função de PDF
        sumario = {
            'titulo': 'Filtros Aplicados:',
            'conteudo': sumario_filtros
        }

        # Chama a função reutilizável para criar o PDF
        pdf_path = criar_pdf_livros(livros_filtrados, 'Relatório de Livros Personalizado', sumario)

        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /livros/relatorio: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório."), 500


@app.route('/livros/relatorio/excel', methods=['GET'])
def gerar_relatorio_livros_excel():
    try:
        cur = con.cursor()

        # A lógica para montar o sumário e buscar os dados é EXATAMENTE a mesma
        # da sua rota de PDF '/livros/relatorio'
        # ... (copie e cole aqui toda a lógica de filtros e sumário) ...
        sumario_filtros = "..."  # Sua lógica para montar esta string

        # Busca os livros filtrados (mesma lógica da rota de PDF)
        sql_base = "SELECT A.titulo, A.autor, A.idioma, A.ano_publicado, A.qtd_disponivel, A.qtd_total, A.isbn FROM ACERVO A"
        # Assumindo que você tem uma função construir_clausulas_filtro como na sua rota de PDF
        where_clause, params = construir_clausulas_filtro(request.args)
        sql_final = f"{sql_base} {where_clause} ORDER BY A.titulo"
        cur.execute(sql_final, tuple(params))
        livros_filtrados = cur.fetchall()
        cur.close()

        # Prepara as informações para a função de Excel
        sumario = {
            'titulo': 'Filtros Aplicados:',
            'conteudo': sumario_filtros
        }

        # --- AQUI ESTÁ A ÚNICA MUDANÇA REAL ---
        # Chama a NOVA função para criar o EXCEL
        excel_path = criar_excel_livros(livros_filtrados, 'Relatório de Livros Personalizado', sumario)

        # Envia o arquivo para o usuário com o mimetype correto
        return send_file(
            excel_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERRO NA ROTA /livros/relatorio/excel: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório em Excel."), 500


# ROTA DE RELATÓRIO COMPLETO (AGORA USA A MESMA LÓGICA DE LAYOUT)
@app.route('/livros/relatorio/completo', methods=['GET'])
def gerar_relatorio_completo():
    try:
        cur = con.cursor()

        # Busca TODOS os livros
        sql = "SELECT titulo, autor, idioma, ano_publicado, qtd_disponivel, qtd_total, isbn FROM ACERVO ORDER BY titulo"
        cur.execute(sql)
        todos_os_livros = cur.fetchall()
        cur.close()

        # Prepara as informações para a função de PDF
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f'Total de {len(todos_os_livros)} livros no acervo.'
        }

        # Chama a MESMA função reutilizável para criar o PDF
        pdf_path = criar_pdf_livros(todos_os_livros, 'Relatório Geral do Acervo', sumario)

        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /livros/relatorio/completo: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório geral."), 500


@app.route('/livros/relatorio/completo/excel', methods=['GET'])
def gerar_relatorio_completo_excel():
    try:
        cur = con.cursor()

        # A busca de dados é IDÊNTICA à da rota de PDF completo
        sql = "SELECT titulo, autor, idioma, ano_publicado, qtd_disponivel, qtd_total, isbn FROM ACERVO ORDER BY titulo"
        cur.execute(sql)
        todos_os_livros = cur.fetchall()
        cur.close()

        # Prepara as informações para a função de Excel
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f'Total de {len(todos_os_livros)} livros no acervo.'
        }

        # --- AQUI ESTÁ A MUDANÇA ---
        # Chama a função para criar o EXCEL
        excel_path = criar_excel_livros(todos_os_livros, 'Relatório Geral do Acervo', sumario)

        # Envia o arquivo para o usuário com o mimetype correto
        return send_file(
            excel_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERRO NA ROTA /livros/relatorio/completo/excel: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório geral em Excel."), 500


# --- NOVA ROTA 2: Para gerar o relatório de livros emprestados ---
@app.route('/livros/relatorio/emprestados', methods=['GET'])
def gerar_relatorio_emprestados():
    """Gera um PDF contendo apenas os livros que estão emprestados."""
    try:
        cur = con.cursor()

        # SQL que busca os detalhes dos livros que estão na lista de emprestados.
        # Usamos uma subquery para isso:
        sql = """
            SELECT A.titulo, A.autor, A.idioma, A.ano_publicado, 
                   A.qtd_disponivel, A.qtd_total, A.isbn
            FROM ACERVO A
            WHERE A.id_livro IN (SELECT ID_LIVRO FROM RESERVAS WHERE STATUS = 2)
            ORDER BY A.titulo
        """
        cur.execute(sql)
        livros_emprestados = cur.fetchall()
        cur.close()

        # Prepara as informações para a nossa função de PDF reutilizável
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f'Total de {len(livros_emprestados)} livros atualmente emprestados.'
        }

        # Chama a MESMA função que já usamos para os outros relatórios!
        # Isso garante que o layout será idêntico e consistente.
        pdf_path = criar_pdf_livros(livros_emprestados, 'Relatório de Livros Emprestados', sumario)

        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /livros/relatorio/emprestados: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório de livros emprestados."), 500


# =========================================================================
#  NOVA ROTA PARA GERAR O RELATÓRIO DE LIVROS EMPRESTADOS EM EXCEL
# =========================================================================
@app.route('/livros/relatorio/emprestados/excel', methods=['GET'])
def gerar_relatorio_emprestados_excel():
    try:
        cur = con.cursor()

        # A busca de dados é IDÊNTICA à da rota de PDF de emprestados
        sql = """
            SELECT A.titulo, A.autor, A.idioma, A.ano_publicado, 
                   A.qtd_disponivel, A.qtd_total, A.isbn
            FROM ACERVO A
            WHERE A.id_livro IN (SELECT ID_LIVRO FROM RESERVAS WHERE STATUS = 2)
            ORDER BY A.titulo
        """
        cur.execute(sql)
        livros_emprestados = cur.fetchall()
        cur.close()

        # Prepara as informações para a função de Excel
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f'Total de {len(livros_emprestados)} livros atualmente emprestados.'
        }

        # --- AQUI ESTÁ A MUDANÇA ---
        # Chama a função para criar o EXCEL
        excel_path = criar_excel_livros(livros_emprestados, 'Relatório de Livros Emprestados', sumario)

        # Envia o arquivo para o usuário
        return send_file(
            excel_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERRO NA ROTA /livros/relatorio/emprestados/excel: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório de emprestados em Excel."), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ESTANTE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= VE TODAS AS ESTANTE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/estantes', methods=['GET'])
def listar_estantes():
    try:
        cursor = con.cursor()
        cursor.execute("SELECT ID_ESTANTE, TITULO, SUBTITULO, IDS_LIVROS FROM ESTANTE")

        estantes = []
        for row in cursor.fetchall():
            estantes.append({
                'id_estante': row[0],
                'titulo': row[1],
                'subtitulo': row[2],
                'ids_livros': list(map(int, row[3].split(','))) if row[3] else []
            })

        return jsonify({'estantes': estantes}), 200

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= VE ESTANTES PELO ID =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/estante/<int:id_estante>', methods=['GET'])
def obter_estante(id_estante):
    try:
        cursor = con.cursor()
        cursor.execute("SELECT ID_ESTANTE, TITULO, SUBTITULO, IDS_LIVROS FROM ESTANTE WHERE ID_ESTANTE = ?",
                       (id_estante,))
        row = cursor.fetchone()

        if not row:
            return jsonify({'erro': 'Estante não encontrada.'}), 404

        estante = {
            'id_estante': row[0],
            'titulo': row[1],
            'subtitulo': row[2],
            'ids_livros': list(map(int, row[3].split(','))) if row[3] else []
        }
        return jsonify(estante), 200

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= CRIA ESTANTE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/adicionar-estante', methods=['POST'])
def adicionar_estante():
    dados = request.get_json()

    titulo = dados.get('titulo')
    subtitulo = dados.get('subtitulo')
    ids_livros = dados.get('ids_livros')

    if not titulo or not isinstance(ids_livros, list):
        return jsonify({'erro': 'Dados inválidos. Envie um título e uma lista de IDs.'}), 400

    try:
        cursor = con.cursor()

        # Verifica se os livros existem no acervo
        cursor.execute(
            "SELECT id_livro FROM ACERVO WHERE id_livro IN ({})".format(
                ','.join('?' for _ in ids_livros)
            ), tuple(ids_livros)
        )
        livros_existentes = {row[0] for row in cursor.fetchall()}

        ids_invalidos = [id_livro for id_livro in ids_livros if id_livro not in livros_existentes]

        if ids_invalidos:
            return jsonify({
                'erro': 'Os seguintes IDs de livros não existem no acervo:',
                'ids_invalidos': ids_invalidos
            }), 400

        # Insere na tabela ESTANTE
        ids_str = ','.join(map(str, ids_livros))

        cursor.execute("""
            INSERT INTO ESTANTE (TITULO, SUBTITULO, IDS_LIVROS)
            VALUES (?, ?, ?)
        """, (titulo, subtitulo, ids_str))

        # Pega o ID da estante recém-criada
        cursor.execute("SELECT MAX(ID_ESTANTE) FROM ESTANTE")
        id_estante = cursor.fetchone()[0]

        # Insere na tabela Layout com padrão inativo
        cursor.execute("""
            INSERT INTO LAYOUT (POSICAO, TIPO, ID_TIPO, ATIVO)
            VALUES (?, ?, ?, ?)
        """, (0, 1, id_estante, 0))  # 1 = tipo estante, ativo = 0 (falso)

        con.commit()

        return jsonify({
            'mensagem': 'Estante adicionada com sucesso!',
            'dados_inseridos': {
                'titulo': titulo,
                'subtitulo': subtitulo,
                'ids_livros': ids_livros,
                'layout': {
                    'tipo': 1,
                    'id_tipo': id_estante,
                    'posicao': 0,
                    'ativo': False
                }
            }
        }), 201

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDIÇÃO DA ESTANTE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/editar-estante/<int:id_estante>', methods=['PUT'])
def editar_estante(id_estante):
    dados = request.get_json()

    titulo = dados.get('titulo')
    subtitulo = dados.get('subtitulo')
    ids_livros = dados.get('ids_livros')

    if not titulo or not isinstance(ids_livros, list):
        return jsonify({'erro': 'Dados inválidos. Envie um título e uma lista de IDs.'}), 400

    try:
        cursor = con.cursor()

        # Validação: checar se os IDs existem no acervo
        cursor.execute(
            "SELECT id_livro FROM ACERVO WHERE id_livro IN ({})".format(
                ','.join('?' for _ in ids_livros)
            ), tuple(ids_livros)
        )
        livros_existentes = {row[0] for row in cursor.fetchall()}
        ids_invalidos = [id_livro for id_livro in ids_livros if id_livro not in livros_existentes]

        if ids_invalidos:
            return jsonify({
                'erro': 'Os seguintes IDs de livros não existem no acervo:',
                'ids_invalidos': ids_invalidos
            }), 400

        # Atualiza a estante
        ids_str = ','.join(map(str, ids_livros))
        cursor.execute("""
            UPDATE ESTANTE
            SET TITULO = ?, SUBTITULO = ?, IDS_LIVROS = ?
            WHERE ID_ESTANTE = ?
        """, (titulo, subtitulo, ids_str, id_estante))

        if cursor.rowcount == 0:
            return jsonify({'erro': 'Estante não encontrada.'}), 404

        con.commit()

        return jsonify({
            'mensagem': 'Estante atualizada com sucesso!',
            'dados_atualizados': {
                'id_estante': id_estante,
                'titulo': titulo,
                'subtitulo': subtitulo,
                'ids_livros': ids_livros
            }
        }), 200

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EXCLÓI ESTANTE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/deletar-estante/<int:id_estante>', methods=['DELETE'])
def deletar_estante(id_estante):
    try:
        cursor = con.cursor()
        cursor.execute("""
            DELETE FROM ESTANTE
            WHERE ID_ESTANTE = ?
        """, (id_estante,))

        if cursor.rowcount == 0:
            return jsonify({'erro': 'Estante não encontrada.'}), 404

        con.commit()

        return jsonify({
            'mensagem': 'Estante deletada com sucesso!',
            'id_estante_deletado': id_estante
        }), 200

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= BANERS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= RECUPERA BANER =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/baners', methods=['GET'])
def listar_baners():
    cursor = con.cursor()
    cursor.execute("SELECT ID_BANER, descricao FROM Baners")
    resultados = cursor.fetchall()

    lista_baners = [{'id_baner': linha[0], 'descricao': linha[1]} for linha in resultados]

    return jsonify(lista_baners), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ADICIONA BANER =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/baners', methods=['POST'])
def criar_baner():
    imagem = request.files.get('imagem')
    descricao = request.form.get('descricao', '').strip()

    if not imagem:
        return jsonify({'erro': 'Imagem obrigatória'}), 400

    if not descricao:
        return jsonify({'erro': 'Descrição obrigatória'}), 400

    try:
        cursor = con.cursor()

        # Verifica se há menos de 10 banners
        cursor.execute("SELECT COUNT(*) FROM Baners")
        total = cursor.fetchone()[0]
        if total >= 10:
            return jsonify({'erro': 'Limite de 10 banners atingido'}), 400

        # Insere o novo banner com descrição e retorna o ID
        cursor.execute("INSERT INTO Baners (descricao) VALUES (?) RETURNING ID_BANER", (descricao,))
        id_baner = cursor.fetchone()[0]

        # Adiciona na tabela LAYOUT
        cursor.execute("""
            INSERT INTO Layout (POSICAO, TIPO, ID_TIPO, ATIVO)
            VALUES (?, ?, ?, ?)
        """, (0, 2, id_baner, 0))  # tipo 2 = banner, ativo = 0 (inativo)

        con.commit()

        # Salva a imagem com nome baseado no ID
        nome_arquivo = f"{id_baner}.jpeg"
        pasta_destino = os.path.join(app.config['UPLOAD_FOLDER'], "Baners")
        os.makedirs(pasta_destino, exist_ok=True)
        caminho_imagem = os.path.join(pasta_destino, nome_arquivo)
        imagem.save(caminho_imagem)

        return jsonify({
            'mensagem': 'Banner cadastrado com sucesso!',
            'id_baner': id_baner,
            'descricao': descricao,
            'layout': {
                'tipo': 2,
                'id_tipo': id_baner,
                'posicao': 0,
                'ativo': False
            }
        }), 201

    except Exception as e:
        print("Erro ao criar banner:", e)
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITA BANER =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/baners/<int:id_baner>', methods=['PUT'])
def editar_baner(id_baner):
    descricao = request.form.get('descricao')
    imagem = request.files.get('imagem')

    cursor = con.cursor()

    # Verifica se o baner existe
    cursor.execute("SELECT 1 FROM Baners WHERE ID_BANER = ?", (id_baner,))
    if not cursor.fetchone():
        return jsonify({"erro": "Baner não encontrado"}), 404

    # Atualiza a descrição
    cursor.execute("UPDATE Baners SET descricao = ? WHERE ID_BANER = ?", (descricao, id_baner))
    con.commit()

    # Se uma nova imagem for enviada, substitui a anterior
    if imagem:
        nome_arquivo = f"{id_baner}.jpeg"
        pasta_destino = os.path.join(app.config['UPLOAD_FOLDER'], "Baners")
        os.makedirs(pasta_destino, exist_ok=True)
        caminho_imagem = os.path.join(pasta_destino, nome_arquivo)
        imagem.save(caminho_imagem)

    return jsonify({"message": "Baner atualizado com sucesso!"}), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EXCLOI BANER =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/baners/<int:id_baner>', methods=['DELETE'])
def excluir_baner(id_baner):
    cursor = con.cursor()

    # Verifica se o banner existe
    cursor.execute("SELECT 1 FROM Baners WHERE ID_BANER = ?", (id_baner,))
    if not cursor.fetchone():
        return jsonify({"erro": "Baner não encontrado"}), 404

    # Deleta o banner do banco
    cursor.execute("DELETE FROM Baners WHERE ID_BANER = ?", (id_baner,))
    con.commit()

    # Tenta remover a imagem associada
    nome_arquivo = f"{id_baner}.jpeg"
    caminho_imagem = os.path.join(app.config['UPLOAD_FOLDER'], "Baners", nome_arquivo)
    if os.path.exists(caminho_imagem):
        os.remove(caminho_imagem)

    return jsonify({"message": "Baner excluído com sucesso!"}), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= COMENTARIOS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= MOSTRA COMENTARIOS DE ACORDO COM O LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/comentarios/livro/<int:id_livro>', methods=['GET'])
def listar_comentarios_por_livro(id_livro):
    cursor = con.cursor()

    # Verifica se o livro existe
    cursor.execute("SELECT 1 FROM ACERVO WHERE ID_LIVRO = ?", (id_livro,))
    if not cursor.fetchone():
        return jsonify({'erro': 'Livro não encontrado'}), 404

    # Busca todos os comentários relacionados ao livro
    cursor.execute("""
        SELECT ID_COMENTARIO, ID_USUARIO, SPOILER, COMENTARIO
        FROM COMENTARIO
        WHERE ID_LIVRO = ?
        ORDER BY ID_COMENTARIO DESC
    """, (id_livro,))

    comentarios = cursor.fetchall()
    resultado = []

    for row in comentarios:
        id_comentario, id_usuario, spoiler, texto = row

        # Busca o nome do usuário
        cursor.execute("SELECT NOME FROM USUARIOS WHERE ID_USUARIO = ?", (id_usuario,))
        usuario = cursor.fetchone()
        nome_usuario = usuario[0] if usuario else "Usuário Desconhecido"

        resultado.append({
            'id_comentario': id_comentario,
            'id_usuario': id_usuario,
            'nome_usuario': nome_usuario,
            'spoiler': bool(spoiler),
            'comentario': texto
        })

    return jsonify({'comentarios': resultado}), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= ADICIONA COMENTARIOS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/comentarios', methods=['POST'])
def criar_comentario():
    data = request.get_json()

    if not data:
        return jsonify({'erro': 'JSON inválido ou não enviado'}), 400

    id_usuario = data.get('id_usuario')
    id_livro = data.get('id_livro')
    spoiler = data.get('spoiler', False)
    comentario = data.get('comentario', '').strip()

    # Validações básicas
    if not id_usuario:
        return jsonify({'erro': 'ID do usuário é obrigatório'}), 400
    if not id_livro:
        return jsonify({'erro': 'ID do livro é obrigatório'}), 400
    if not comentario:
        return jsonify({'erro': 'Comentário não pode estar vazio'}), 400

    cursor = con.cursor()

    # Verifica se o usuário existe
    cursor.execute("SELECT 1 FROM USUARIOS WHERE ID_USUARIO = ?", (id_usuario,))
    if not cursor.fetchone():
        return jsonify({'erro': 'Usuário não encontrado'}), 404

    # Verifica se o livro existe
    cursor.execute("SELECT 1 FROM ACERVO WHERE ID_LIVRO = ?", (id_livro,))
    if not cursor.fetchone():
        return jsonify({'erro': 'Livro não encontrado'}), 404

    # Insere o comentário
    cursor.execute("""
        INSERT INTO COMENTARIO (ID_USUARIO, ID_LIVRO, SPOILER, COMENTARIO)
        VALUES (?, ?, ?, ?)
        RETURNING ID_COMENTARIO
    """, (id_usuario, id_livro, spoiler, comentario))

    id_comentario = cursor.fetchone()[0]
    con.commit()

    return jsonify({
        'mensagem': 'Comentário cadastrado com sucesso!',
        'id_comentario': id_comentario,
        'id_usuario': id_usuario,
        'id_livro': id_livro,
        'spoiler': spoiler
    }), 201


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITA COMENTARIOS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/comentarios/<int:id_comentario>', methods=['PUT'])
def atualizar_spoiler_comentario(id_comentario):
    data = request.get_json()

    if not data:
        return jsonify({'erro': 'JSON inválido ou não enviado'}), 400

    # Permite apenas o campo "spoiler"
    if 'spoiler' not in data or len(data.keys()) > 1:
        return jsonify({'erro': 'Só é permitido atualizar o campo "spoiler"'}), 400

    spoiler = data.get('spoiler')

    if spoiler is None:
        return jsonify({'erro': 'Campo "spoiler" inválido'}), 400

    cursor = con.cursor()

    # Verifica se o comentário existe
    cursor.execute("SELECT 1 FROM COMENTARIO WHERE ID_COMENTARIO = ?", (id_comentario,))
    if not cursor.fetchone():
        return jsonify({'erro': 'Comentário não encontrado'}), 404

    # Atualiza apenas o campo "spoiler"
    cursor.execute("UPDATE COMENTARIO SET SPOILER = ? WHERE ID_COMENTARIO = ?", (bool(spoiler), id_comentario))
    con.commit()

    return jsonify({'mensagem': 'Campo "spoiler" atualizado com sucesso!'}), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EXCLUIR COMENTARIOS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/comentarios/<int:id_comentario>', methods=['DELETE'])
def deletar_comentario(id_comentario):
    cursor = con.cursor()

    # Verifica se o comentário existe
    cursor.execute("SELECT 1 FROM COMENTARIO WHERE ID_COMENTARIO = ?", (id_comentario,))
    if not cursor.fetchone():
        return jsonify({'erro': 'Comentário não encontrado'}), 404

    # Exclui o comentário
    cursor.execute("DELETE FROM COMENTARIO WHERE ID_COMENTARIO = ?", (id_comentario,))
    con.commit()

    return jsonify({'mensagem': 'Comentário excluído com sucesso!'}), 200


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= AGENDAMENTO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= LISTAR AGENDAMENTOS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/agendamentos', methods=['GET'])
def listar_agendamentos():
    try:
        cur = con.cursor()

        # Query SQL atualizada para incluir o valor da multa e a data de vencimento original
        cur.execute("""
            SELECT 
                r.id_reserva,
                u.nome AS nome_usuario,
                a.titulo AS titulo_livro,
                r.data_reservado,
                r.data_validade,
                r.status,
                r.data_devolver,      -- Prazo de devolução quando emprestado
                r.data_devolvido,     -- Data real da devolução
                m.STATUS AS status_multa,
                m.id_multa,
                u.id_usuario,
                a.id_livro,
                m.Valor_total AS valor_multa,  -- ADICIONADO: Valor total da multa
                r.data_devolver AS data_vencimento_original -- ADICIONADO: Campo que o frontend precisa
            FROM RESERVAS r
            JOIN USUARIOS u ON r.id_usuario = u.id_usuario
            JOIN ACERVO a ON r.id_livro = a.id_livro
            LEFT JOIN MULTAS m ON r.id_reserva = m.id_reserva
            ORDER BY r.data_reservado DESC
        """)

        reservas = cur.fetchall()
        cur.close()

        if not reservas:
            return jsonify({"message": "Nenhum agendamento encontrado."}), 404

        agendamentos = []
        for reserva in reservas:
            agendamento = {
                "id_reserva": reserva[0],
                "nome_usuario": reserva[1],
                "titulo_livro": reserva[2],
                "data_reservado": reserva[3],
                "data_validade": reserva[4],
                "status": reserva[5],
                "data_devolver": reserva[6],
                "data_devolvido": reserva[7],
                "status_multa": reserva[8] if reserva[8] is not None else 0,
                "id_multa": reserva[9],
                "id_usuario": reserva[10],
                "id_livro": reserva[11],
                # ADICIONADO: Mapeando os novos campos para o JSON
                "valor_multa": float(reserva[12]) if reserva[12] is not None else 0.0,
                "data_vencimento_original": reserva[13]
            }
            agendamentos.append(agendamento)

        return jsonify(agendamentos), 200

    except Exception as e:
        return jsonify({"message": f"Erro ao listar agendamentos: {str(e)}"}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= HISTORICO (DO USUARIO) DE AGENDAMENTO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/agendamentos/<int:id_usuario>', methods=['GET'])
def listar_agendamentos_usuario(id_usuario):
    try:
        cur = con.cursor()
        cur.execute("""
            SELECT r.id_reserva, r.data_reservado, r.data_validade, r.status,
                   a.id_livro, a.titulo, a.autor, a.categoria, a.isbn, 
                   a.qtd_disponivel, a.sinopse, a.idioma, a.ano_publicado
            FROM RESERVAS r
            JOIN ACERVO a ON r.id_livro = a.id_livro
            WHERE r.id_usuario = ?
            ORDER BY r.data_reservado DESC
        """, (id_usuario,))
        reservas = cur.fetchall()
        cur.close()

        if not reservas:
            return jsonify({"message": "Nenhuma reserva encontrada para este usuário."}), 404

        agora = datetime.now()
        agendamentos = []
        for reserva in reservas:
            data_validade = reserva[2]
            status = reserva[3]
            multado = data_validade < agora and status not in (3, 4)

            agendamento = {
                "id_reserva": reserva[0],
                "data_reservado": reserva[1].strftime('%Y-%m-%d %H:%M:%S'),
                "data_validade": data_validade.strftime('%Y-%m-%d %H:%M:%S'),
                "status": status,
                "multado": multado,
                "livro": {
                    "id": reserva[4],
                    "titulo": reserva[5],
                    "autor": reserva[6],
                    "categoria": reserva[7],
                    "isbn": reserva[8],
                    "qtd_disponivel": reserva[9],
                    "sinopse": reserva[10],
                    "idioma": reserva[11],
                    "ano_publicado": reserva[12]
                }
            }
            agendamentos.append(agendamento)

        return jsonify(agendamentos), 200

    except Exception as e:
        return jsonify({"message": f"Erro ao buscar agendamentos: {str(e)}"}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PESQUISA AGENDAMENTO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/agendamento/<int:id>', methods=['GET'])
def agendamentos(id):
    try:
        cur = con.cursor()
        cur.execute("""
            SELECT RESERVAS.id_reserva,
                   RESERVAS.id_usuario,
                   RESERVAS.id_livro,
                   ACERVO.titulo,
                   USUARIOS.nome,
                   RESERVAS.data_reservado,
                   RESERVAS.data_validade,
                   RESERVAS.status,
                   MULTAS.status
            FROM RESERVAS
            LEFT JOIN ACERVO ON ACERVO.id_livro = RESERVAS.id_livro
            LEFT JOIN USUARIOS ON USUARIOS.id_usuario = RESERVAS.id_usuario
            LEFT JOIN MULTAS ON MULTAS.id_reserva = RESERVAS.id_reserva
            where reservas.id_reserva = ?
        """, (id,))
        reservas = cur.fetchall()

        reserva_dic = []
        for reserva in reservas:
            reserva_dic.append({
                'id_reserva': reserva[0],
                'id_usuario': reserva[1],
                'id_livro': reserva[2],
                'titulo': reserva[3],
                'nome': reserva[4],
                'data_reservado': reserva[5],
                'data_validade': reserva[6],
                'status': reserva[7],
                'status_multa': reserva[8]
            })

        return jsonify(mensagem='Reservas Feitas', reservas=reserva_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PESQUISA AGENDAMENTO PELO ID DO LIVRO OU USUARIO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/verificar-reserva', methods=['GET'])
def verificar_reserva_expecifica():
    id_usuario = request.args.get('id_usuario')
    id_livro = request.args.get('id_livro')

    if not id_usuario and not id_livro:
        return jsonify(mensagem="Parâmetros id_usuario ou id_livro são necessários"), 400

    conditions = []
    params = []

    if id_usuario:
        id_usuario = int(id_usuario.strip())
        conditions.append("RESERVAS.id_usuario = ?")
        params.append(id_usuario)

    if id_livro:
        id_livro = int(id_livro.strip())
        conditions.append("RESERVAS.id_livro = ?")
        params.append(id_livro)

    where_clause = " OR ".join(conditions)
    query = f"""
        SELECT FIRST 1
               RESERVAS.id_reserva,
               RESERVAS.id_usuario,
               RESERVAS.id_livro,
               ACERVO.titulo,
               USUARIOS.nome,
               RESERVAS.data_reservado,
               RESERVAS.data_validade,
               RESERVAS.status
        FROM RESERVAS
        LEFT JOIN ACERVO ON ACERVO.id_livro = RESERVAS.id_livro
        LEFT JOIN USUARIOS ON USUARIOS.id_usuario = RESERVAS.id_usuario
        WHERE {where_clause}
        ORDER BY RESERVAS.data_reservado DESC
    """

    try:
        cur = con.cursor()
        cur.execute(query, tuple(params))
        reserva = cur.fetchone()

        if not reserva:
            return jsonify(mensagem="Nenhuma reserva encontrada"), 404

        data_reservado = reserva[5].strftime('%Y-%m-%d %H:%M:%S')
        data_validade = reserva[6].strftime('%Y-%m-%d %H:%M:%S')
        reserva_dic = {
            'id_reserva': reserva[0],
            'id_usuario': reserva[1],
            'id_livro': reserva[2],
            'titulo': reserva[3],
            'nome': reserva[4],
            'data_reservado': data_reservado,
            'data_validade': data_validade,
            'status': reserva[7]
        }

        return jsonify(mensagem='Reserva encontrada', reserva=reserva_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= VERIFICADOR DE STATUS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route("/api/verificar-reserva")
def verificar_reserva():
    id_usuario = request.args.get("id_usuario")
    id_livro = request.args.get("id_livro")

    cur = con.cursor()
    cur.execute("""
        SELECT STATUS
        FROM RESERVAS
        WHERE ID_USUARIO = ? AND ID_LIVRO = ?
        ORDER BY DATA_RESERVADO DESC
        FETCH FIRST 1 ROWS ONLY
    """, (id_usuario, id_livro))
    reserva = cur.fetchone()
    cur.close()

    if reserva:
        return jsonify({"status": reserva[0]})
    else:
        return jsonify({"status": 0})


@app.route('/estatisticas/movimentacao-status-por-dia', methods=['GET'])
def estatisticas_movimentacao_status():
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if not start_date_str or not end_date_str:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=29)
            end_date_str = end_date.strftime('%Y-%m-%d')
            start_date_str = start_date.strftime('%Y-%m-%d')

        query_firebird = """
            WITH RECURSIVE DateSeries (DIA) AS (
                SELECT CAST(? AS DATE) FROM RDB$DATABASE
                UNION ALL
                SELECT DATEADD(1 DAY TO d.DIA)
                FROM DateSeries d
                WHERE d.DIA < CAST(? AS DATE)
            ),
            Eventos AS (
                SELECT CAST(R.DATA_RESERVADO AS DATE) AS DIA_EVENTO, 'RESERVADO' AS TIPO_EVENTO
                FROM RESERVAS R
                WHERE R.STATUS >= 1 AND R.DATA_RESERVADO BETWEEN ? AND ?
                UNION ALL
                SELECT CAST(R.DATA_VALIDADE AS DATE) AS DIA_EVENTO, 'EMPRESTADO' AS TIPO_EVENTO
                FROM RESERVAS R
                WHERE R.STATUS >= 2 AND R.DATA_VALIDADE BETWEEN ? AND ?
                UNION ALL
                SELECT CAST(R.DATA_DEVOLVIDO AS DATE) AS DIA_EVENTO, 'DEVOLVIDO' AS TIPO_EVENTO
                FROM RESERVAS R
                WHERE R.STATUS = 4 AND R.DATA_DEVOLVIDO BETWEEN ? AND ?
            ),
            MovimentacoesAgrupadas AS (
                SELECT
                    E.DIA_EVENTO,
                    CAST(SUM(CASE WHEN E.TIPO_EVENTO = 'RESERVADO' THEN 1 ELSE 0 END) AS INTEGER) AS TOTAL_RESERVADOS,
                    CAST(SUM(CASE WHEN E.TIPO_EVENTO = 'EMPRESTADO' THEN 1 ELSE 0 END) AS INTEGER) AS TOTAL_EMPRESTADOS,
                    CAST(SUM(CASE WHEN E.TIPO_EVENTO = 'DEVOLVIDO' THEN 1 ELSE 0 END) AS INTEGER) AS TOTAL_DEVOLVIDOS
                FROM Eventos E
                GROUP BY E.DIA_EVENTO
            )
            SELECT
                DS.DIA,
                COALESCE(M.TOTAL_RESERVADOS, 0) AS TOTAL_RESERVADOS,
                COALESCE(M.TOTAL_EMPRESTADOS, 0) AS TOTAL_EMPRESTADOS,
                COALESCE(M.TOTAL_DEVOLVIDOS, 0) AS TOTAL_DEVOLVIDOS
            FROM DateSeries DS
            LEFT JOIN MovimentacoesAgrupadas M ON DS.DIA = M.DIA_EVENTO
            ORDER BY DS.DIA
        """

        cur = con.cursor()

        # ==========================================================
        # AQUI ESTÁ A CORREÇÃO: A tupla agora tem 8 elementos
        # ==========================================================
        params = (
            start_date_str, end_date_str,  # Para a DateSeries (2)
            start_date_str, end_date_str,  # Para os eventos de Reserva (2)
            start_date_str, end_date_str,  # Para os eventos de Empréstimo (2)
            start_date_str, end_date_str  # Para os eventos de Devolução (2)
        )
        cur.execute(query_firebird, params)

        dados_brutos = cur.fetchall()
        cur.close()

        # O resto do código permanece idêntico
        labels = []
        dados_reservados = []
        dados_emprestados = []
        dados_devolvidos = []

        for registro in dados_brutos:
            if registro[0] is not None:
                labels.append(registro[0].strftime('%d/%m/%Y'))
                dados_reservados.append(int(registro[1] or 0))
                dados_emprestados.append(int(registro[2] or 0))
                dados_devolvidos.append(int(registro[3] or 0))

        resposta_final = {
            "labels": labels,
            "datasets": [
                {"label": "Reservados", "data": dados_reservados},
                {"label": "Emprestados", "data": dados_emprestados},
                {"label": "Devolvidos", "data": dados_devolvidos}
            ]
        }

        return jsonify(resposta_final), 200

    except Exception as e:
        print(f"ERRO NO BACKEND: {e}")
        return jsonify({"message": f"Erro ao gerar estatísticas: {str(e)}"}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= FAZER AGENDAMENTO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/agendamento', methods=['POST'])
def criar_agendamento():
    try:
        dados = request.get_json()
        id_usuario = dados.get('id_usuario')
        id_livro = dados.get('id_livro')

        if not dados or 'id_usuario' not in dados or 'id_livro' not in dados:
            return jsonify({"message": "Dados incompletos ou inválidos"}), 400

        cur = con.cursor()
        cur.execute("""SELECT QTD_DISPONIVEL FROM ACERVO WHERE ID_LIVRO = ?""", (id_livro,))
        tem_livro = cur.fetchone()
        cur.close()

        if not tem_livro or tem_livro[0] == 0:
            return jsonify({"message": "Este livro não está disponível!"}), 400

        # Verificar se o usuário já tem 3 agendamentos ativos (status diferentes de 3 e 4)
        cur = con.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM RESERVAS 
            WHERE ID_USUARIO = ? AND STATUS NOT IN (3, 4)
        """, (id_usuario,))
        qtd_agendamentos = cur.fetchone()[0]
        cur.close()

        # Pega o maxiomo de livros que podem ser emprestados
        cur = con.cursor()
        cur.execute("""
            SELECT MAX_LIVROS FROM VALORES
            ORDER BY ID_VALORES DESC
            ROWS 1
        """)
        max_livro = cur.fetchone()[0]
        cur.close()

        print(max_livro)

        if qtd_agendamentos >= max_livro:
            return jsonify({"message": f"Você só pode agendar {max_livro} livros por vez!"}), 400

        # Verificar se o usuário tem alguma multa pendente (status 1)
        cur = con.cursor()
        cur.execute("""
            SELECT 1 FROM MULTAS 
            WHERE ID_USUARIO = ? AND STATUS = 1
        """, (id_usuario,))
        multa_pendente = cur.fetchone()
        cur.close()

        if multa_pendente:
            return jsonify(
                {"message": "Você possui multas pendentes! Regularize antes de fazer novos agendamentos."}), 403

        cur = con.cursor()
        cur.execute("""SELECT 1 FROM RESERVAS WHERE STATUS in(1,2) AND ID_USUARIO = ? AND ID_LIVRO = ?""",
                    (id_usuario, id_livro))
        tem_reserva = cur.fetchone()
        cur.close()

        if tem_reserva:
            return jsonify({"message": "Já existe um agendamento em andamento para este usuário!"}), 400

        agora = datetime.now()
        data_reservado = agora.strftime('%Y-%m-%d %H:%M:%S')

        hora_fechamento = HORARIO_FUNC_FIM
        fechamento_hoje = datetime.combine(agora.date(), hora_fechamento)

        if agora > fechamento_hoje:
            dia_retirada = agora.date() + timedelta(days=1)
        else:
            dia_retirada = agora.date()

        data_validade = datetime.combine(dia_retirada, hora_fechamento)
        data_validade_str = data_validade.strftime('%Y-%m-%d %H:%M:%S')
        data_reservado_obj = datetime.strptime(data_reservado, '%Y-%m-%d %H:%M:%S')

        # Criar reserva
        cur = con.cursor()
        cur.execute(""" 
            INSERT INTO RESERVAS (id_usuario, id_livro, data_reservado, data_validade, status)
            VALUES (?, ?, ?, ?, ?) RETURNING id_reserva
        """, (id_usuario, id_livro, data_reservado, data_validade_str, 1))
        id_reserva = cur.fetchone()[0]
        con.commit()

        # Atualizar quantidade disponível do livro
        cur.execute(""" 
            UPDATE ACERVO SET QTD_DISPONIVEL = QTD_DISPONIVEL - 1
            WHERE id_livro = ?
        """, (id_livro,))
        con.commit()

        # Buscar informações do usuário para o e-mail
        cur.execute("SELECT email, nome FROM USUARIOS WHERE id_usuario = ?", (id_usuario,))
        usuario = cur.fetchone()
        email_usuario = usuario[0]
        nome_usuario = usuario[1]

        # Buscar título do livro
        cur.execute("SELECT titulo, autor FROM ACERVO WHERE id_livro = ?", (id_livro,))
        livro = cur.fetchone()
        titulo_livro = livro[0]
        autor_livro = livro[1]

        # Enviar e-mail
        corpo_email = f"""
        <p>Olá <strong>{nome_usuario}</strong>, seu agendamento foi criado com sucesso!</p>

        <p>Segue abaixo as informações do agendamento:</p>

        <p>
            <strong>Título:</strong> {titulo_livro}<br>
            <strong>Autor:</strong> {autor_livro}<br>

            <strong>Data da Reserva:</strong> {data_reservado_obj.strftime('%d/%m/%Y às %H:%M')}<br>

            <strong>Validade para Retirada:</strong> dia {data_validade.strftime('%d/%m/%Y')}, das {HORARIO_FUNC_INICIO} às {HORARIO_ALMOCO_INICIO} e das {HORARIO_ALMOCO_FIM} às {HORARIO_FUNC_FIM}<br>
            <strong>Observação:</strong> Caso a retirada não seja feita no prazo, o agendamento terá que ser feito novamente.
        </p>

        <p><strong>📍 Local da Retirada:</strong><br>
            {RUA}, {NUMERO} - {BAIRRO}, {CIDADE}<br>
            <a href="https://www.google.com/maps/place/Biblioteca+Municipal/@-21.2911253,-50.3426846,17z/data=!3m1!4b1!4m6!3m5!1s0x949614d7c9ef99a3:0xd3d8d11cd629e99!8m2!3d-21.2911253!4d-50.3403334!16s%2Fg%2F11cm6fwmq6?entry=ttu" style="color: #1a73e8;">Clique aqui para ver no Google Maps</a>
        </p>

        <h3>📖 Orientações para Retirada:</h3>
        <ul>
            <li>Traga um documento com foto.</li>
            <li>Apresente este e-mail na recepção.</li>
            <li>Chegue até o horário indicado para evitar cancelamento automático.</li>
            <li>Em caso de dúvidas, entre em contato com a biblioteca.</li>
        </ul>

        <p>Obrigado por utilizar nosso sistema! 📚</p>
        """

        enviar_email_com_template(
            destinatario=email_usuario,
            assunto="📚 Agendamento Confirmado!",
            titulo="Agendamento Confirmado",
            corpo=corpo_email
        )

        data_formatada = data_validade.strftime('%d/%m/%Y')
        return jsonify(
            mensagem=f"Reserva criada com sucesso! Você pode retirar o livro das {HORARIO_FUNC_INICIO} às {HORARIO_ALMOCO_INICIO} e das {HORARIO_ALMOCO_FIM} às {HORARIO_FUNC_FIM} do dia {data_formatada}."), 201

    except Exception as e:
        return jsonify(mensagem=f"Erro ao criar reserva: {str(e)}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EXPIRAÇÃO AUTOMATICA DO AGENDAMENTO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
def atualizar_status_agendamento():
    try:
        agora = datetime.now()
        cur = con.cursor()
        # Seleciona reservas expiradas (status 1 e data_validade vencida)
        cur.execute("""
            SELECT id_reserva, id_livro
            FROM RESERVAS
            WHERE status = 1 AND data_validade < ?
        """, (agora.strftime('%Y-%m-%d %H:%M:%S'),))

        reservas_para_atualizar = cur.fetchall()

        for reserva in reservas_para_atualizar:
            id_reserva = reserva[0]
            id_livro = reserva[1]

            # Atualiza o status da reserva para expirada (3)
            cur.execute("""
                UPDATE RESERVAS
                SET status = 3
                WHERE id_reserva = ?
            """, (id_reserva,))

            # Atualiza a quantidade disponível do livro (+1)
            cur.execute("""
                UPDATE ACERVO
                SET qtd_disponivel = qtd_disponivel + 1
                WHERE id_livro = ?
            """, (id_livro,))

        con.commit()
        cur.close()
        return "Reservas expiradas atualizadas e livros devolvidos ao acervo."

    except Exception as e:
        return f"Erro ao atualizar status: {e}"


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITAR AGENDAMENTO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/agendamento/status/<int:id>', methods=['PUT'])
def editar_agendamento(id):
    try:
        dados = request.get_json()
        status = dados.get('status')

        cur = con.cursor()

        # Buscar status atual, data_validade, data_devolver e id_livro da reserva
        cur.execute("""
            SELECT status, DATA_VALIDADE, DATA_RESERVADO, DATA_DEVOLVER, id_livro, id_usuario 
            FROM RESERVAS 
            WHERE id_reserva = ?
        """, (id,))
        reserva = cur.fetchone()

        if not reserva:
            return jsonify({'mensagem': 'Reserva não encontrada!'}), 404

        # Desempacotamento direto dos dados
        status_atual, data_validade, data_reservado, _, id_livro, id_usuario = reserva

        # =-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-= FAZER O EMPRESTIMO =-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=
        if status_atual == 1 and status == 2:
            if data_validade:
                if isinstance(data_validade, datetime):
                    data_devolucao = data_validade + timedelta(days=15)
                else:
                    data_validade = datetime.strptime(data_validade, "%Y-%m-%d")
                    data_devolucao = data_validade + timedelta(days=15)

                cur.execute("""
                    UPDATE RESERVAS
                    SET status = ?, DATA_DEVOLVER = ?
                    WHERE id_reserva = ?
                """, (status, data_devolucao.strftime("%Y-%m-%d"), id))
                con.commit()

                # ---------------- email de cancelamento ----------------------------------------

                if not isinstance(data_reservado, datetime):
                    data_reservado_obj = datetime.strptime(data_reservado, '%Y-%m-%d %H:%M:%S')
                else:
                    data_reservado_obj = data_reservado

                cur.execute("SELECT email, nome FROM USUARIOS WHERE id_usuario = ?", (id_usuario,))
                usuario = cur.fetchone()
                email_usuario = usuario[0]
                nome_usuario = usuario[1]

                # Buscar título do livro
                cur.execute("SELECT titulo, autor FROM ACERVO WHERE id_livro = ?", (id_livro,))
                livro = cur.fetchone()
                titulo_livro = livro[0]
                autor_livro = livro[1]

                corpo_email = f"""
                    <p>Olá, <strong>{nome_usuario}</strong>!</p>

                    <p>Informamos que o livro reservado foi <strong>retirado com sucesso</strong>.</p>

                    <p>Fique atento à data de devolução para evitar multas e manter o bom uso da biblioteca por todos.</p>

                    <p><strong>📘 Detalhes da retirada:</strong></p>
                    <p>
                      <strong>Título:</strong> {titulo_livro}<br>
                      <strong>Autor:</strong> {autor_livro}<br>
                      <strong>Data da Reserva:</strong> {data_reservado_obj.strftime('%d/%m/%Y às %H:%M')}<br>
                      <strong>Data para Devolução:</strong> {data_devolucao.strftime("%d/%m/%Y")}<br>
                    </p>

                    <p><strong>📍 Local para devolução:</strong><br>
                      {RUA}, {NUMERO} - {BAIRRO}, {CIDADE}<br>
                      <a href="https://www.google.com/maps/place/Biblioteca+Municipal/@-21.2911253,-50.3426846,17z/data=!3m1!4b1!4m6!3m5!1s0x949614d7c9ef99a3:0xd3d8d11cd629e99!8m2!3d-21.2911253!4d-50.3403334!16s%2Fg%2F11cm6fwmq6?entry=ttu" style="color: #1a73e8;">Clique aqui para ver no Google Maps</a>
                    </p>

                    <p>Aproveite a leitura e lembre-se de devolver o livro dentro do prazo! 📚</p>
                """

                enviar_email_com_template(
                    destinatario=email_usuario,
                    assunto="📚 Livro Retirado!",
                    titulo="Livro Retirado",
                    corpo=corpo_email
                )

            else:
                return jsonify({'mensagem': 'Data de validade não encontrada!'}), 400


        # =-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-= DEVOLVER =-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=
        elif status_atual == 2 and status == 4:

            data_devolvido = datetime.now()  # Mantém datetime para comparação
            data_devolvido_str = data_devolvido.strftime("%Y-%m-%d")

            cur.execute("""
                UPDATE RESERVAS
                SET status = ?, DATA_DEVOLVIDO = ?
                WHERE id_reserva = ?
            """, (status, data_devolvido_str, id))
            con.commit()

            # Atualiza quantidade disponível do livro no acervo
            cur.execute("""
                UPDATE ACERVO
                SET qtd_disponivel = qtd_disponivel + 1
                WHERE id_livro = ?
            """, (id_livro,))
            con.commit()

            # Agora buscar DATA_DEVOLVER e ID_USUARIO para verificar atraso
            cur.execute("""
                SELECT ID_USUARIO, DATA_DEVOLVER
                FROM RESERVAS
                WHERE id_reserva = ?
            """, (id,))
            resultado = cur.fetchone()

            if resultado:
                id_usuario, data_devolver = resultado
                # Converter data_devolver para datetime, se necessário
                if isinstance(data_devolver, str):
                    data_devolver = datetime.strptime(data_devolver, "%Y-%m-%d")

                # Verificar se houve atraso
                if data_devolvido > data_devolver:
                    # Chamar aplicar_multa, passando os parâmetros necessários
                    aplicar_multa(id, id_usuario, data_devolver, data_devolvido)
                else:
                    # ---------------- email de devolução ----------------------------------------

                    if not isinstance(data_reservado, datetime):
                        data_reservado_obj = datetime.strptime(data_reservado, '%Y-%m-%d %H:%M:%S')
                    else:
                        data_reservado_obj = data_reservado

                    cur.execute("SELECT email, nome FROM USUARIOS WHERE id_usuario = ?", (id_usuario,))
                    usuario = cur.fetchone()
                    email_usuario = usuario[0]
                    nome_usuario = usuario[1]

                    # Buscar título do livro
                    cur.execute("SELECT titulo, autor FROM ACERVO WHERE id_livro = ?", (id_livro,))
                    livro = cur.fetchone()
                    titulo_livro = livro[0]
                    autor_livro = livro[1]

                    corpo_email = f"""
                        <p>Olá, <strong>{nome_usuario}</strong>!</p>

                        <p>Informamos que o livro que estava em sua posse foi <strong>devolvido com sucesso</strong>.</p>

                        <p>Agradecemos por utilizar a nossa biblioteca e por cumprir com os prazos estipulados. Seu compromisso ajuda a manter o acervo sempre disponível para todos os leitores.</p>

                        <p><strong>📘 Detalhes da devolução:</strong></p>
                        <p>
                          <strong>Título:</strong> {titulo_livro}<br>
                          <strong>Autor:</strong> {autor_livro}<br>
                          <strong>Data da Reserva:</strong> {data_reservado_obj.strftime('%d/%m/%Y às %H:%M')}<br>
                          <strong>Data da Devolução:</strong> {data_devolvido.strftime('%d/%m/%Y')}<br>
                        </p>

                        <p>Esperamos vê-lo novamente em breve. Boa leitura e até a próxima!</p>
                    """

                    enviar_email_com_template(
                        destinatario=email_usuario,
                        assunto="📚 Livro Devolvido!",
                        titulo="Livro Devolvido",
                        corpo=corpo_email
                    )
            else:
                print("Reserva não encontrada para verificar atraso")


        # =-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-= CANCELAR =-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=-+-=
        elif status_atual == 1 and status == 3:
            # Atualiza status para expirada
            cur.execute("""
                UPDATE RESERVAS
                SET status = ?
                WHERE id_reserva = ?
            """, (status, id))
            con.commit()

            # Aumenta o livro devolvido no acervo
            cur.execute("""
                UPDATE ACERVO
                SET qtd_disponivel = qtd_disponivel + 1
                WHERE id_livro = ?
            """, (id_livro,))
            con.commit()

            if not isinstance(data_reservado, datetime):
                data_reservado_obj = datetime.strptime(data_reservado, '%Y-%m-%d %H:%M:%S')
            else:
                data_reservado_obj = data_reservado

            # ---------------- email de cancelamento ----------------------------------------
            cur.execute("SELECT email, nome FROM USUARIOS WHERE id_usuario = ?", (id_usuario,))
            usuario = cur.fetchone()
            email_usuario = usuario[0]
            nome_usuario = usuario[1]

            # Buscar título do livro
            cur.execute("SELECT titulo, autor FROM ACERVO WHERE id_livro = ?", (id_livro,))
            livro = cur.fetchone()
            titulo_livro = livro[0]
            autor_livro = livro[1]

            corpo_email = f"""
                    <p>Olá, <strong>{nome_usuario}</strong>!</p>

                    <p>Informamos que sua reserva foi <strong>cancelada</strong>.</p>

                    <p>Esse cancelamento pode ter ocorrido por um dos seguintes motivos:</p>
                    <ul>
                      <li>O livro não foi retirado dentro do prazo estipulado;</li>
                      <li>A reserva foi cancelada manualmente por você;</li>
                      <li>O bibliotecário realizou o cancelamento.</li>
                    </ul>

                    <p>Se nenhuma das opções acima se aplica ao seu caso, por favor, entre em contato conosco para mais informações.</p>

                    <p><strong>📘 Detalhes da reserva:</strong></p>
                    <p>
                      <strong>Título:</strong> {titulo_livro}<br>
                      <strong>Autor:</strong> {autor_livro}<br>
                      <strong>Data da Reserva:</strong> {data_reservado_obj.strftime('%d/%m/%Y às %H:%M')}<br>
                      <strong>Validade para Retirada:</strong> dia {data_validade.strftime('%d/%m/%Y')}, das {HORARIO_FUNC_INICIO} às {HORARIO_ALMOCO_INICIO} e das {HORARIO_ALMOCO_FIM} às {HORARIO_FUNC_FIM}<br>
                      <strong>Observação:</strong> Após o prazo, será necessário realizar uma nova reserva.
                    </p>

                    <p><strong>📍 Local para retirada:</strong><br>
                      {RUA}, {NUMERO} - {BAIRRO}, {CIDADE}<br>
                      <a href="https://www.google.com/maps/place/Biblioteca+Municipal/@-21.2911253,-50.3426846,17z/data=!3m1!4b1!4m6!3m5!1s0x949614d7c9ef99a3:0xd3d8d11cd629e99!8m2!3d-21.2911253!4d-50.3403334!16s%2Fg%2F11cm6fwmq6?entry=ttu" style="color: #1a73e8;">Clique aqui para ver no Google Maps</a>
                    </p>

                    <p>Agradecemos por utilizar nosso sistema. Até a próxima! 📚</p>

                    """

            enviar_email_com_template(
                destinatario=email_usuario,
                assunto="📚 Agendamento Cancelado!",
                titulo="Agendamento Cancelado",
                corpo=corpo_email
            )


        else:
            # Apenas atualiza o status
            cur.execute("""
                UPDATE RESERVAS
                SET status = ?
                WHERE id_reserva = ?
            """, (status, id))

        con.commit()

        return jsonify({
            'message': "Status atualizado com sucesso!",
            'livro': {
                'id_reserva': id,
                'status': status
            }
        })

    except Exception as e:
        return jsonify({'mensagem': f"Erro ao atualizar status: {e}"}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PDF DOS AGENDAMENTOS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
# =========================================================================
#  1. SUBSTITUA A SUA FUNÇÃO DE GERAR O PDF POR ESTA VERSÃO ATUALIZADA
# =========================================================================
def criar_pdf_movimentacao_detalhada(dados_por_tipo, titulo_relatorio, sumario_info, colunas_pdf):
    # 'P' para Retrato (Vertical)
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)

    # --- FUNÇÃO INTERNA PARA REPETIR O CABEÇALHO EM NOVAS PÁGINAS ---
    def desenhar_cabecalho_pagina():
        pdf.add_page()
        try:
            pdf.image('static/uploads/logo.png', x=10, y=8, w=30)
        except FileNotFoundError:
            pass  # Ignora se o logo não existir

        pdf.set_font('Arial', 'B', 18)
        pdf.cell(0, 10, titulo_relatorio, ln=True, align='C')
        pdf.set_font('Arial', '', 8)
        data_hoje = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        pdf.cell(0, 5, f"Gerado em: {data_hoje}", ln=True, align='C')
        pdf.ln(5)

        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 7, sumario_info['titulo'], ln=True)
        pdf.set_font('Arial', '', 9)
        pdf.multi_cell(0, 5, sumario_info['conteudo'], border=1, align='L')
        pdf.ln(5)

    # --- INÍCIO DO DOCUMENTO ---
    desenhar_cabecalho_pagina()

    # --- DESENHO DAS TABELAS (UMA PARA CADA TIPO DE MOVIMENTAÇÃO) ---
    for tipo, lista_de_dados in dados_por_tipo.items():
        # Verifica se a tabela cabe na página atual, se não, cria uma nova
        if pdf.get_y() > 220:  # Deixa uma margem segura no final da página
            desenhar_cabecalho_pagina()

        # Título da tabela (ex: "Tabela de Movimentações: Reservas")
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, f"Tabela de Movimentações: {tipo}", ln=True, align='L')

        # Se não houver dados para este tipo, informa o usuário
        if not lista_de_dados:
            pdf.set_font('Arial', 'I', 10)
            pdf.cell(0, 10, "Nenhuma movimentação encontrada para este tipo no período.", ln=True)
            pdf.ln(5)
            continue  # Pula para o próximo tipo de relatório

        # Cabeçalho da tabela
        pdf.set_font('Arial', 'B', 9)
        pdf.set_fill_color(230, 230, 230)
        for titulo, largura in colunas_pdf:
            pdf.cell(w=largura, h=7, txt=titulo, border=1, align='C', fill=True)
        pdf.ln()

        # Linhas de dados
        pdf.set_font('Arial', '', 8)
        for registro in lista_de_dados:
            # Quebra de página se a próxima linha não couber
            if pdf.get_y() > 270:
                pdf.add_page()
                pdf.ln(10)  # Espaço no topo
                pdf.set_font('Arial', 'B', 9)
                for titulo, largura in colunas_pdf:
                    pdf.cell(w=largura, h=7, txt=titulo, border=1, align='C', fill=True)
                pdf.ln()
                pdf.set_font('Arial', '', 8)

            # Prepara e codifica os dados da linha para evitar erros
            dados_linha = []
            for item in registro:
                if isinstance(item, datetime):
                    dados_linha.append(item.strftime('%d/%m/%Y'))
                else:
                    dados_linha.append(str(item))

            dados_enc = [d.encode('latin-1', 'ignore').decode('latin-1') for d in dados_linha]

            for i, dado in enumerate(dados_enc):
                pdf.cell(w=colunas_pdf[i][1], h=6, txt=dado, border=1, align='L')
            pdf.ln()
        pdf.ln(10)  # Espaço entre as tabelas

    # Salva o arquivo e retorna o caminho
    pdf_path = f"relatorio_detalhado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(pdf_path)
    return pdf_path


# =========================================================================
#  2. SUBSTITUA A SUA ROTA POR ESTA VERSÃO COMPLETA E CORRIGIDA
# =========================================================================
@app.route('/relatorio/movimentacao/detalhado', methods=['GET'])
def gerar_relatorio_movimentacao_detalhado():
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        tipos_str = request.args.get('tipos')

        if not all([start_date_str, end_date_str, tipos_str]):
            return jsonify(mensagem="Parâmetros de data e tipos são obrigatórios."), 400

        tipos_solicitados = tipos_str.split(',')
        dados_para_pdf = {}
        cur = con.cursor()

        colunas_pdf = [("Usuário", 50), ("Título do Livro", 70), ("Data do Evento", 30), ("Status/Prazo", 40)]

        # --- SQL para RESERVAS (CORRIGIDO) ---
        if 'reservados' in tipos_solicitados:
            sql_reservados = """
                SELECT U.NOME, A.TITULO, R.DATA_RESERVADO,
                       'Reservado neste dia'
                FROM RESERVAS R
                JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
                JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
                WHERE R.STATUS >= 1 AND CAST(R.DATA_RESERVADO AS DATE) BETWEEN ? AND ?
                ORDER BY R.DATA_RESERVADO
            """
            cur.execute(sql_reservados, (start_date_str, end_date_str))
            dados_para_pdf['Reservas'] = cur.fetchall()
            print(f"Encontrados {len(dados_para_pdf['Reservas'])} registros para Reservas.")  # LINHA DE DEBUG

        # --- SQL para EMPRÉSTIMOS (CORRIGIDO) ---
        if 'emprestados' in tipos_solicitados:
            sql_emprestados = """
                SELECT U.NOME, A.TITULO, R.DATA_VALIDADE,
                       'Devolver até: ' || CAST(R.DATA_DEVOLVER AS VARCHAR(10))
                FROM RESERVAS R
                JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
                JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
                WHERE R.STATUS >= 2 AND CAST(R.DATA_VALIDADE AS DATE) BETWEEN ? AND ?
                ORDER BY R.DATA_VALIDADE
            """
            cur.execute(sql_emprestados, (start_date_str, end_date_str))
            dados_para_pdf['Empréstimos'] = cur.fetchall()
            print(f"Encontrados {len(dados_para_pdf['Empréstimos'])} registros para Empréstimos.")  # LINHA DE DEBUG

        # --- SQL para DEVOLUÇÕES (CORRIGIDO) ---
        if 'devolvidos' in tipos_solicitados:
            sql_devolvidos = """
                SELECT U.NOME, A.TITULO, R.DATA_DEVOLVIDO,
                       'Devolvido neste dia'
                FROM RESERVAS R
                JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
                JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
                WHERE R.STATUS = 4 AND CAST(R.DATA_DEVOLVIDO AS DATE) BETWEEN ? AND ?
                ORDER BY R.DATA_DEVOLVIDO
            """
            cur.execute(sql_devolvidos, (start_date_str, end_date_str))
            dados_para_pdf['Devoluções'] = cur.fetchall()
            print(f"Encontrados {len(dados_para_pdf['Devoluções'])} registros para Devoluções.")  # LINHA DE DEBUG

        cur.close()

        # O resto do código permanece o mesmo...
        data_inicio_fmt = datetime.strptime(start_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        data_fim_fmt = datetime.strptime(end_date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        sumario = {
            'titulo': 'Filtros Aplicados:',
            'conteudo': f"Período: De {data_inicio_fmt} a {data_fim_fmt} | Tipos: {', '.join(tipos_solicitados).replace('-', ' ').title()}"
        }

        pdf_path = criar_pdf_movimentacao_detalhada(dados_para_pdf, 'Relatório Detalhado de Movimentação', sumario,
                                                    colunas_pdf)
        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/movimentacao/detalhado: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório detalhado."), 500


# ==================================================================================
#  ROTA 1: RELATÓRIO DE RESERVAS AGUARDANDO RETIRADA (STATUS = 1)
# ==================================================================================
@app.route('/relatorio/reservas/aguardando-retirada', methods=['GET'])
def gerar_relatorio_aguardando_retirada():
    try:
        cur = con.cursor()
        # SQL que busca APENAS status 1
        sql = """
            SELECT 
                U.NOME, 
                A.TITULO,
                R.DATA_RESERVADO,
                R.DATA_VALIDADE
            FROM RESERVAS R
            JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
            JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
            WHERE R.STATUS = 1
            ORDER BY R.DATA_RESERVADO DESC
        """
        cur.execute(sql)
        reservas_aguardando = cur.fetchall()
        cur.close()

        dados_para_pdf = {'Reservas Aguardando Retirada': reservas_aguardando}
        colunas_pdf = [("Usuário", 50), ("Título do Livro", 70), ("Data da Reserva", 30), ("Prazo", 40)]
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f"Total de {len(reservas_aguardando)} livros aguardando retirada."
        }

        pdf_path = criar_pdf_movimentacao_detalhada(dados_para_pdf, 'Relatório de Reservas Ativas', sumario,
                                                    colunas_pdf)
        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/reservas/aguardando-retirada: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório."), 500


# =========================================================================
#  NOVA ROTA PARA GERAR O RELATÓRIO DE RESERVAS AGUARDANDO RETIRADA EM EXCEL
# =========================================================================
@app.route('/relatorio/reservas/aguardando-retirada/excel', methods=['GET'])
def gerar_relatorio_aguardando_retirada_excel():
    try:
        cur = con.cursor()

        # A busca de dados é IDÊNTICA à da rota de PDF
        sql = """
            SELECT 
                U.NOME, 
                A.TITULO,
                R.DATA_RESERVADO,
                R.DATA_VALIDADE
            FROM RESERVAS R
            JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
            JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
            WHERE R.STATUS = 1
            ORDER BY R.DATA_RESERVADO DESC
        """
        cur.execute(sql)
        reservas_aguardando = cur.fetchall()
        cur.close()

        # Para manter a consistência, vamos criar um dicionário de dados
        dados_para_excel = {
            'Reservas Aguardando Retirada': reservas_aguardando
        }

        # Prepara o sumário
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f"Total de {len(reservas_aguardando)} livros aguardando retirada."
        }

        # --- AQUI ESTÁ A MUDANÇA ---
        # Chama uma função genérica para criar o Excel (vamos precisar dela)
        # Assumindo que temos uma função como criar_excel_generico
        excel_path = criar_excel_generico(
            dados_para_excel,
            'Relatório de Reservas Ativas',
            sumario,
            ["Usuário", "Título do Livro", "Data da Reserva", "Válido até"]
        )

        # Envia o arquivo para o usuário
        return send_file(
            excel_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/reservas/aguardando-retirada/excel: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório em Excel."), 500


def criar_excel_generico(dados_por_tipo, titulo_relatorio, sumario_info, lista_de_colunas):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_relatorio[:30]  # Limita o nome da aba

    # Estilos (pode copiar da sua outra função de excel)
    font_titulo = Font(name='Calibri', size=18, bold=True, color='2F5496')
    alignment_center = Alignment(horizontal='center', vertical='center')
    font_cabecalho_tabela = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_cabecalho_tabela = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Cabeçalho do Arquivo
    ws.merge_cells('A1:D1')
    ws['A1'] = titulo_relatorio
    ws['A1'].font = font_titulo
    ws['A1'].alignment = alignment_center

    # Sumário
    ws.merge_cells('A3:D3')
    ws['A3'] = sumario_info['titulo']
    ws.merge_cells('A4:D4')
    ws['A4'] = sumario_info['conteudo']

    linha_atual = 6
    for tipo, lista_de_dados in dados_por_tipo.items():
        ws[f'A{linha_atual}'] = tipo
        ws[f'A{linha_atual}'].font = Font(bold=True, size=12)
        linha_atual += 1

        # Cabeçalho da Tabela
        ws.append(lista_de_colunas)
        for col in range(1, len(lista_de_colunas) + 1):
            cell = ws.cell(row=linha_atual, column=col)
            cell.font = font_cabecalho_tabela
            cell.fill = fill_cabecalho_tabela
        linha_atual += 1

        # Dados
        for registro in lista_de_dados:
            dados_linha = list(registro)
            # Formata datas, se houver
            for i, item in enumerate(dados_linha):
                if isinstance(item, datetime):
                    dados_linha[i] = item.strftime('%d/%m/%Y')
            ws.append(dados_linha)

        linha_atual += 2

    # Ajuste de colunas
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column_letter].width = max_length + 2

    excel_path = f"relatorio_generico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(excel_path)
    return excel_path


# ==================================================================================
#  ROTA 2: RELATÓRIO DE EMPRÉSTIMOS ATUAIS (STATUS = 2)
# ==================================================================================
@app.route('/relatorio/emprestimos/atuais', methods=['GET'])
def gerar_relatorio_emprestimos_atuais():
    try:
        cur = con.cursor()
        # SQL que busca APENAS status 2
        sql = """
            SELECT 
                U.NOME, 
                A.TITULO,
                R.DATA_VALIDADE,
                'Devolver até: ' || CAST(CAST(R.DATA_DEVOLVER AS DATE) AS VARCHAR(10))
            FROM RESERVAS R
            JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
            JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
            WHERE R.STATUS = 2
            ORDER BY R.DATA_DEVOLVER ASC
        """
        cur.execute(sql)
        emprestimos_atuais = cur.fetchall()
        cur.close()

        dados_para_pdf = {'Empréstimos Atuais': emprestimos_atuais}
        colunas_pdf = [("Usuário", 50), ("Título do Livro", 70), ("Data da Retirada", 30), ("Prazo de Devolução", 40)]
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f"Total de {len(emprestimos_atuais)} livros atualmente emprestados."
        }

        pdf_path = criar_pdf_movimentacao_detalhada(dados_para_pdf, 'Relatório de Empréstimos Atuais', sumario,
                                                    colunas_pdf)
        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/emprestimos/atuais: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório."), 500


# =========================================================================
#  NOVA ROTA PARA GERAR O RELATÓRIO DE EMPRÉSTIMOS ATUAIS EM EXCEL
# =========================================================================
@app.route('/relatorio/emprestimos/atuais/excel', methods=['GET'])
def gerar_relatorio_emprestimos_atuais_excel():
    try:
        cur = con.cursor()

        # A busca de dados é IDÊNTICA à da rota de PDF
        sql = """
            SELECT 
                U.NOME, 
                A.TITULO,
                R.DATA_VALIDADE,
                'Devolver até: ' || CAST(CAST(R.DATA_DEVOLVER AS DATE) AS VARCHAR(10))
            FROM RESERVAS R
            JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
            JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
            WHERE R.STATUS = 2
            ORDER BY R.DATA_DEVOLVER ASC
        """
        cur.execute(sql)
        emprestimos_atuais = cur.fetchall()
        cur.close()

        # Prepara os dados para o Excel
        dados_para_excel = {
            'Empréstimos Atuais': emprestimos_atuais
        }

        # Prepara o sumário
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f"Total de {len(emprestimos_atuais)} livros atualmente emprestados."
        }

        # --- AQUI ESTÁ A MUDANÇA ---
        # Chama a função genérica para criar o Excel
        excel_path = criar_excel_generico(
            dados_para_excel,
            'Relatório de Empréstimos Atuais',
            sumario,
            ["Usuário", "Título do Livro", "Data da Retirada", "Prazo de Devolução"]
        )

        # Envia o arquivo para o usuário
        return send_file(
            excel_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/emprestimos/atuais/excel: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório em Excel."), 500


@app.route('/relatorio/multas/pendentes', methods=['GET'])
def gerar_relatorio_multas_pendentes():
    try:
        cur = con.cursor()

        # Consulta SQL com COALESCE para segurança
        sql = """
            SELECT 
                U.NOME, 
                A.TITULO,
                R.DATA_DEVOLVER,
                COALESCE(M.VALOR_TOTAL, 0) AS VALOR_TOTAL -- Garante que nunca seja NULL
            FROM MULTAS M
            JOIN RESERVAS R ON M.ID_RESERVA = R.ID_RESERVA
            JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
            JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
            WHERE M.STATUS = 1
            ORDER BY R.DATA_DEVOLVER ASC
        """

        cur.execute(sql)
        multas_pendentes = cur.fetchall()
        cur.close()

        # --- AQUI ACONTECE A TRANSFORMAÇÃO ---
        # 1. Calcula o valor total antes de formatar
        valor_total_arrecadar = sum(multa[3] for multa in multas_pendentes)

        # 2. Cria uma nova lista com os valores já formatados como string
        dados_formatados = []
        for nome, titulo, data_devolver, valor in multas_pendentes:
            # Formata o valor da multa como moeda brasileira (ex: 1.087,50)
            valor_formatado = locale.currency(valor, grouping=True)
            dados_formatados.append((nome, titulo, data_devolver, valor_formatado))

        # 3. Prepara os dados para a função de PDF, agora usando a lista formatada
        dados_para_pdf = {
            'Multas Pendentes de Pagamento': dados_formatados  # <--- Usa a lista formatada
        }

        # Define as colunas (o cabeçalho do valor agora indica que é R$)
        colunas_pdf = [("Usuário Devedor", 60), ("Título do Livro", 70), ("Devolução", 30),
                       ("Valor", 30)]  # Aumentei um pouco a largura

        # Prepara o sumário com o valor total também formatado
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f"Total de {len(multas_pendentes)} multas pendentes. Valor total a receber: {locale.currency(valor_total_arrecadar, grouping=True)}"
        }

        # Chama a função de PDF
        pdf_path = criar_pdf_movimentacao_detalhada(dados_para_pdf, 'Relatório de Multas Pendentes', sumario,
                                                    colunas_pdf)

        return send_file(pdf_path, as_attachment=True, mimetype='application/pdf')

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/multas/pendentes: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório de multas."), 500


# =========================================================================
#  NOVA ROTA PARA GERAR O RELATÓRIO DE MULTAS PENDENTES EM EXCEL
# =========================================================================
@app.route('/relatorio/multas/pendentes/excel', methods=['GET'])
def gerar_relatorio_multas_pendentes_excel():
    try:
        cur = con.cursor()

        # A busca de dados é IDÊNTICA à da rota de PDF
        sql = """
            SELECT 
                U.NOME, 
                A.TITULO,
                R.DATA_DEVOLVER,
                COALESCE(M.VALOR_TOTAL, 0) AS VALOR_TOTAL
            FROM MULTAS M
            JOIN RESERVAS R ON M.ID_RESERVA = R.ID_RESERVA
            JOIN USUARIOS U ON R.ID_USUARIO = U.ID_USUARIO
            JOIN ACERVO A ON R.ID_LIVRO = A.ID_LIVRO
            WHERE M.STATUS = 1
            ORDER BY R.DATA_DEVOLVER ASC
        """
        cur.execute(sql)
        multas_pendentes = cur.fetchall()  # Dados brutos com números
        cur.close()

        # Prepara os dados para o Excel (usando os números brutos)
        dados_para_excel = {
            'Multas Pendentes de Pagamento': multas_pendentes
        }

        # Calcula e formata o valor total para o sumário
        valor_total_arrecadar = sum(multa[3] for multa in multas_pendentes)
        sumario = {
            'titulo': 'Resumo do Relatório:',
            'conteudo': f"Total de {len(multas_pendentes)} multas pendentes. Valor total a receber: {locale.currency(valor_total_arrecadar, grouping=True)}"
        }

        # --- AQUI ESTÁ A MUDANÇA ---
        # Chama a função genérica para criar o Excel
        excel_path = criar_excel_generico(
            dados_para_excel,
            'Relatório de Multas Pendentes',
            sumario,
            ["Usuário Devedor", "Título do Livro", "Prazo de Devolução", "Valor a Pagar"]
        )

        # Envia o arquivo para o usuário
        return send_file(
            excel_path,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERRO NA ROTA /relatorio/multas/pendentes/excel: {e}")
        return jsonify(mensagem="Erro ao gerar o relatório de multas em Excel."), 500


# =========================================================================
#  NOVA ROTA PARA OBTER OS TOTAIS DE RESERVAS, EMPRÉSTIMOS E MULTAS
# =========================================================================
@app.route('/estatisticas/totais-ativos', methods=['GET'])
def get_totais_ativos():
    try:
        cur = con.cursor()

        # 1. Conta as reservas ativas (Status = 1)
        cur.execute("SELECT COUNT(ID_RESERVA) FROM RESERVAS WHERE STATUS = 1")
        total_reservas = cur.fetchone()[0]

        # 2. Conta os empréstimos ativos (Status = 2)
        cur.execute("SELECT COUNT(ID_RESERVA) FROM RESERVAS WHERE STATUS = 2")
        total_emprestimos = cur.fetchone()[0]

        # 3. Conta as multas pendentes (Status = 1)
        cur.execute("SELECT COUNT(ID_MULTA) FROM MULTAS WHERE STATUS = 1")
        total_multas = cur.fetchone()[0]

        cur.close()

        # Retorna todos os totais em um único objeto JSON
        return jsonify({
            'total_reservas': total_reservas,
            'total_emprestimos': total_emprestimos,
            'total_multas': total_multas
        })

    except Exception as e:
        print(f"ERRO NA ROTA /estatisticas/totais-ativos: {e}")
        return jsonify(mensagem="Erro ao buscar totais."), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= RESERVA =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= FAZER RESERVA =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/aviseme', methods=['POST'])
def avisar():
    dados = request.get_json()
    email = dados.get('email')
    id_livro = dados.get('id_livro')
    try:
        cur = con.cursor()
        cur.execute("""
                        SELECT DISTINCT 1 FROM AVISEME WHERE id_livro = ? AND email = ? 
                    """, (id_livro, email,))
        tem_reserva = cur.fetchone()
        cur.close()

        if tem_reserva:
            return jsonify({"message": "Já existe uma reserva em andamento para este usupario"})

        cur = con.cursor()
        cur.execute("""
                INSERT INTO AVISEME (email, id_livro)
                VALUES (?, ?) 
            """, (email, id_livro))
        con.commit()

        return jsonify(
            mensagem=f"Você será avisado quando o livro estiver disponivel."), 201

    except Exception as e:
        return jsonify(mensagem=f"Erro ao criar reserva: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= TAG =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= PESQUISAR TAG POR ID =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/tag/<int:id_tag>', methods=['GET'])
def buscar_nome_tag(id_tag):
    try:
        cur = con.cursor()
        cur.execute("SELECT nome_tag FROM TAGS WHERE id_tag = ?", (id_tag,))
        resultado = cur.fetchone()

        if resultado:
            return jsonify(id_tag=id_tag, nome=resultado[0])
        else:
            return jsonify(mensagem="Tag não encontrada"), 404

    except Exception as e:
        return jsonify(mensagem=f"Erro ao buscar a tag: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= CRIAR TAG =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/tags', methods=['POST'])
def adicionar_tag():
    try:
        dados = request.get_json()
        nome_tag = dados.get('nome_tag')  # Alterado aqui também!

        if not nome_tag:
            return jsonify({'erro': 'Nome da tag é obrigatório'}), 400

        cur = con.cursor()

        # Verificar se a tag já existe
        cur.execute("SELECT * FROM TAGS WHERE NOME_TAG = ?", (nome_tag,))
        tag_existente = cur.fetchone()

        if tag_existente:
            return jsonify({'erro': 'Tag já existe'}), 409

        # Inserir nova tag
        cur.execute("INSERT INTO TAGS (NOME_TAG) VALUES (?)", (nome_tag,))
        con.commit()

        return jsonify({'mensagem': 'Tag adicionada com sucesso'}), 201

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= REMOVER TAG DO LIVRO =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/remover_tag', methods=['POST'])
def remover_tag():
    try:
        data = request.get_json()
        id_livro = data['id_livro']
        id_tag = data['id_tag']

        cur = con.cursor()

        # Primeiro, verificar se a associação existe
        cur.execute("SELECT 1 FROM LIVRO_TAGS WHERE id_livro = ? AND id_tag = ?", (id_livro, id_tag))
        association = cur.fetchone()

        if not association:
            return jsonify({'erro': 'Essa tag não está vinculada a este livro.'})

        # Buscar o nome do livro e da tag
        cur.execute("SELECT TITULO FROM ACERVO WHERE id_livro = ?", (id_livro,))
        titulo_livro = cur.fetchone()
        cur.execute("SELECT nome_tag FROM TAGS WHERE id_tag = ?", (id_tag,))
        nome_tag = cur.fetchone()

        titulo_livro = titulo_livro[0] if titulo_livro else "Livro não encontrado"
        nome_tag = nome_tag[0] if nome_tag else "Tag não encontrada"

        # Agora sim, remover
        cur.execute("DELETE FROM LIVRO_TAGS WHERE id_livro = ? AND id_tag = ?", (id_livro, id_tag))
        con.commit()

        return jsonify({'mensagem': f'A tag "{nome_tag}" foi removida do livro "{titulo_livro}".'})
    except Exception as e:
        return jsonify({'erro': str(e)})


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= BUSCA LIVRO POR TAG =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/acervo/tag/<int:id_tag>', methods=['GET'])
def acervo_por_tag(id_tag):
    """
    Busca e retorna todos os livros que possuem uma tag específica.
    O ID da tag é passado como um inteiro na URL.
    """
    try:
        # Criar cursor para consulta
        cur = con.cursor()

        # A consulta SQL agora usa um JOIN para filtrar pela tabela LIVRO_TAGS.
        # Usamos SELECT DISTINCT para garantir que um livro não apareça
        # duplicado caso haja algum erro ou inconsistência nos dados.
        sql_query = """
            SELECT 
                DISTINCT A.id_livro,
                A.titulo,
                A.autor,
                A.categoria,
                A.isbn,
                A.qtd_disponivel,
                A.sinopse,
                A.idioma,
                A.ano_publicado,
                A.qtd_total,
                A.avaliacao,
                A.editora,
                A.paginas,
                A.maior,
                A.local_fisico,
                -- Este subselect é importante para pegar TODAS as tags do livro,
                -- não apenas a que usamos para filtrar.
                (SELECT LIST(lt2.id_tag)
                 FROM LIVRO_TAGS lt2
                 WHERE lt2.id_livro = A.id_livro) AS tags_ids
            FROM ACERVO A
            -- O JOIN é a forma mais eficiente de conectar as tabelas
            JOIN LIVRO_TAGS LT ON A.id_livro = LT.id_livro
            -- E o WHERE filtra para trazer apenas os livros com a tag desejada
            WHERE LT.id_tag = ?
        """

        # Executa a consulta passando o id_tag como parâmetro
        cur.execute(sql_query, (id_tag,))

        livros_filtrados = cur.fetchall()

        if not livros_filtrados:
            return jsonify(mensagem=f"Nenhum livro encontrado com a tag ID: {id_tag}"), 404

        # A partir daqui, o código é EXATAMENTE o mesmo dos endpoints anteriores,
        # garantindo que a resposta JSON seja consistente.
        acervo_dic = []
        for livro in livros_filtrados:
            id_livro = livro[0]
            tags_ids = livro[15]

            imagem_url = url_for('static', filename=f"Uploads/Livros/{id_livro}.jpeg", _external=True)

            acervo_dic.append({
                'id_livro': id_livro,
                'titulo': livro[1],
                'autor': livro[2],
                'categoria': livro[3],
                'isbn': livro[4],
                'qtd_disponivel': livro[5],
                'sinopse': livro[6],
                'idioma': livro[7],
                'ano_publicado': livro[8],
                'qtd_total': livro[9],
                'avaliacao': livro[10],
                'editora': livro[11],
                'paginas': livro[12],
                'maior': livro[13],
                'local_fisico': livro[14],
                'tags_ids': tags_ids,
                'imagens': [imagem_url]
            })

        return jsonify(mensagem=f"Livros encontrados para a tag ID: {id_tag}", acervo=acervo_dic)

    except Exception as e:
        return jsonify(mensagem=f"Erro ao consultar o banco de dados por tag: {e}"), 500


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= MULTAS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= RETORNA TODAS AS MULTAS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/multas', methods=['GET'])
def get_multas():
    cur = con.cursor()
    query = "SELECT * FROM MULTAS"
    cur.execute(query)
    resultados = cur.fetchall()
    return jsonify(resultados)


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= RETORNA AS MULTAS DE UM USUARIO SÓ =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/multas/<int:id_usuario>', methods=['GET'])
def get_multas_por_usuario(id_usuario):
    cur = con.cursor()
    query = """
        SELECT M.id_multa, M.id_usuario, M.id_reserva, M.valor_total, M.status,
               A.id_livro, A.titulo, A.autor, A.ano_publicado
        FROM MULTAS M
        JOIN RESERVAS R ON M.id_reserva = R.id_reserva
        JOIN ACERVO A ON R.id_livro = A.id_livro
        WHERE M.ID_USUARIO = ? AND M.STATUS = 1
    """
    cur.execute(query, (id_usuario,))
    resultados = cur.fetchall()

    multas_formatadas = []
    for row in resultados:
        multa = {
            "id_multa": row[0],
            "id_usuario": row[1],
            "id_reserva": row[2],
            "valor_total": row[3],
            "status": row[4],
            "livro": {
                "id_livro": row[5],
                "titulo": row[6],
                "autor": row[7],
                "ano_publicacao": row[8]
            }
        }
        multas_formatadas.append(multa)

    return jsonify(multas_formatadas)


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= EDITA AS MULTAS =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/multas/<int:id>', methods=['PUT'])
def editar_multa(id):
    cursor = con.cursor()
    dados = request.get_json()

    novo_status = dados.get("status")
    novo_valor = dados.get("valor_total")

    if novo_status is None and novo_valor is None:
        cursor.close()
        return jsonify({"erro": "Nenhum dado para atualizar."}), 400

    campos = []
    valores = []

    if novo_status is not None:
        campos.append("status = ?")
        valores.append(novo_status)
    if novo_valor is not None:
        campos.append("valor_total = ?")
        valores.append(novo_valor)

    valores.append(id)

    query = f"UPDATE MULTAS SET {', '.join(campos)} WHERE id_multa = ?"
    cursor.execute(query, valores)
    con.commit()

    # Busca os dados atualizados da multa
    cursor.execute("SELECT id_multa, status, valor_total FROM MULTAS WHERE id_multa = ?", (id,))
    multa_atualizada = cursor.fetchone()
    cursor.close()

    if not multa_atualizada:
        return jsonify({"erro": "Multa não encontrada após atualização."}), 404

    return jsonify({
        "mensagem": "Multa atualizada com sucesso!",
        "multa": {
            "id_multa": multa_atualizada[0],
            "status": multa_atualizada[1],
            "valor_total": multa_atualizada[2]
        }
    })


# =+-+=+-+=+-+=+-+=+-+=+-+= CALCULA E ADICIONA A MULTA AUTOMATICAMENTE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
def aplicar_multa(id_reserva, id_usuario, data_devolver, data_devolvido):
    try:
        cur = con.cursor()

        if isinstance(data_devolver, str):
            data_devolver = datetime.strptime(data_devolver, "%Y-%m-%d")
        if isinstance(data_devolvido, str):
            data_devolvido = datetime.strptime(data_devolvido, "%Y-%m-%d")

        cur = con.cursor()
        cur.execute("""
            SELECT VALOR_BASE, VALOR_ACRESCIMO FROM VALORES
            ORDER BY ID_VALORES DESC
            ROWS 1
        """)
        valor_base, valor_acrescimo = cur.fetchone()
        cur.close()

        dias_atraso = (data_devolvido - data_devolver).days
        multa_total = valor_base + (valor_acrescimo * dias_atraso)

        cur.execute("""
            INSERT INTO MULTAS (ID_USUARIO, ID_RESERVA, VALOR_TOTAL, STATUS)
            VALUES (?, ?, ?, 1) returning id_multa
        """, (id_usuario, id_reserva, multa_total))
        id_multa = cur.fetchone()[0]
        con.commit()

        # Buscar dados do usuário
        cur.execute("SELECT NOME, EMAIL FROM USUARIOS WHERE ID_USUARIO = ?", (id_usuario,))
        nome_usuario, email_usuario = cur.fetchone()

        # Buscar dados da reserva e livro
        cur.execute("""
            SELECT a.TITULO, a.AUTOR, r.DATA_RESERVADO
            FROM ACERVO a
            JOIN RESERVAS r ON r.ID_LIVRO = a.ID_LIVRO
            WHERE r.ID_RESERVA = ?
        """, (id_reserva,))

        titulo_livro, autor_livro, data_reservado = cur.fetchone()
        nome_arquivo = f"qrcode_pix_{id_multa}.png"

        print("vai entrar para calcular multa")
        imagem_qrcode, caminho_arquivo, payload_completo = gerar_qrcode_pix(multa_total, nome_arquivo)

        # Envio para Imgur
        response = requests.post(
            "https://api.imgur.com/3/image",
            headers={"Authorization": "Client-ID 549d28ca869b1f7"},
            files={"image": open(imagem_qrcode, "rb")},
            data={"type": "image", "title": "Simple upload"}
        )
        image_link = response.json()["data"]["link"] if response.status_code == 200 else None

        corpo_email = f"""
                  <p>Olá <strong>{nome_usuario}</strong>,</p>
                  <p>Verificamos que houve um atraso na devolução do livro <strong>{titulo_livro}</strong>, de autoria de <strong>{autor_livro}</strong>.</p>
                  <p>A devolução estava prevista para <strong>{data_devolver.strftime('%d/%m/%Y')}</strong>, mas foi realizada em <strong>{data_devolvido.strftime('%d/%m/%Y')}</strong>, gerando <strong>{dias_atraso} dia(s)</strong> de atraso.</p>
                  <p>Valor total da multa:</p>
                  <p style="font-size: 18px; font-weight: bold; color: #d9534f;">R$ {multa_total:.2f}</p>
                  <p>O pagamento pode ser feito na biblioteca ou via PIX:</p>
                  <img src="{image_link}" alt="QR Code PIX" width="200"><br>
                  <p><strong>Código PIX:</strong><br>{payload_completo}</p>
                  <h3>📘 Detalhes da Reserva:</h3>
                  <p><strong>Data da Reserva:</strong> {data_reservado.strftime('%d/%m/%Y às %H:%M')}<br>
                  <strong>Data Devolvida:</strong> {data_devolvido.strftime('%d/%m/%Y')}</p>
              """
        enviar_email_com_template(
            destinatario=email_usuario,
            assunto="📚 Multa por Atraso na Devolução",
            titulo="Multa Registrada",
            corpo=corpo_email,
            caminho_imagem=imagem_qrcode
        )

        return jsonify({'mensagem': 'Multa aplicada e e-mail enviado com sucesso!', 'id_multa': id_multa})

    except Exception as e:
        return jsonify({'mensagem': f'Erro ao aplicar multa: {e}'}), 500

    finally:
        cur.close()


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+= QRCODE =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+#
def format_tlv(id, value):
    return f"{id}{len(value):02d}{value}"


def gerar_qrcode_pix(valor, nome_arquivo):
    CIDADE = app.config['CIDADE']
    NOME_FANTASIA = app.config['NOME_FANTASIA']
    PIX = app.config['PIX']

    print("entrou para calcular multa")
    try:
        # Formata o valor recebido para duas casas decimais (ex: 50.0 -> "50.00")
        valor_formatado = f"{float(valor):.2f}"

        # Se o nome estiver presente, trunca para 25 caracteres; senão, usa valor padrão
        print(f"{NOME_FANTASIA} antes de passar pela formatação 25")
        NOME_FANTASIA = NOME_FANTASIA[:25] if NOME_FANTASIA else "Recebedor PIX"
        print(f"{NOME_FANTASIA} depois de passar pela formatação 25")

        # Trunca a cidade para 15 caracteres, com valor padrão caso esteja ausente
        print(f"{CIDADE} antes de passar pela formatação 15")
        CIDADE = CIDADE[:15] if CIDADE else "Cidade"
        print(f"{CIDADE} depois de passar pela formatação 15")

        # Monta as informações da conta do recebedor (formato exigido pelo padrão PIX)
        merchant_account_info = (
                format_tlv("00", "br.gov.bcb.pix") +  # Identificador do domínio PIX
                format_tlv("01", PIX)  # Chave PIX do recebedor
        )

        # Cria o campo 26 com os dados de conta do recebedor (obrigatório no QR PIX)
        campo_26 = format_tlv("26", merchant_account_info)

        # Constrói o payload completo do QR Code (exceto o CRC final)
        payload_sem_crc = (
                "000201" +  # Identificador do formato do QR Code (versão 01)
                "010212" +  # Transação do tipo "dinâmica com valor"
                campo_26 +  # Dados da conta do recebedor
                "52040000" +  # Código de categoria do ponto de venda (fixo)
                "5303986" +  # Moeda: BRL (986 = Real Brasileiro)
                format_tlv("54", valor_formatado) +  # Valor da transação
                "5802BR" +  # País de origem (BR = Brasil)
                format_tlv("59", NOME_FANTASIA) +  # Nome do recebedor
                format_tlv("60", CIDADE) +  # Cidade do recebedor
                format_tlv("62", format_tlv("05", "***")) +  # Informações adicionais, como referência
                "6304"  # Campo reservado para o CRC16 (será adicionado a seguir)
        )

        # Calcula o CRC16 do payload para garantir integridade do código
        crc = calcula_crc16(payload_sem_crc)

        # Finaliza o payload do QR Code incluindo o CRC calculado
        payload_completo = payload_sem_crc + crc

        # Inicializa o gerador de QR Code com configurações de qualidade
        qr_obj = qrcode.QRCode(
            version=None,  # Deixa a biblioteca definir o tamanho adequado automaticamente
            error_correction=ERROR_CORRECT_H,  # Alta correção de erro (recomendado para pagamentos)
            box_size=10,  # Tamanho dos blocos do QR Code
            border=4  # Tamanho da borda branca em torno do código
        )

        # Adiciona os dados gerados ao QR Code
        qr_obj.add_data(payload_completo)

        # Finaliza a construção do QR Code ajustando ao conteúdo
        qr_obj.make(fit=True)

        # Gera a imagem do QR Code com cores padrão (preto sobre branco)
        qr = qr_obj.make_image(fill_color="black", back_color="white")

        # Define o caminho onde o QR Code será salvo
        pasta_qrcodes = os.path.join(os.getcwd(), "uploads", "qrcodes")

        # Garante que o diretório existe (cria se necessário)
        os.makedirs(pasta_qrcodes, exist_ok=True)

        # Se não for fornecido um nome de arquivo, pode-se gerar um automaticamente (comentado aqui)
        if not nome_arquivo:
            arquivos_existentes = [f for f in os.listdir(pasta_qrcodes) if f.startswith("pix_") and f.endswith(".png")]
            numeros_usados = []
            for nome_arq in arquivos_existentes:
                try:
                    num = int(nome_arq.replace("pix_", "").replace(".png", ""))
                    numeros_usados.append(num)
                except ValueError:
                    continue

        # Define o caminho final do arquivo de imagem do QR Code
        caminho_arquivo = os.path.join(pasta_qrcodes, nome_arquivo)

        # Salva o QR Code como imagem PNG
        qr.save(caminho_arquivo)

        # (Opcional) imprime o payload no console, útil para depuração

        # Retorna o caminho do arquivo e o nome gerado/salvo
        return caminho_arquivo, nome_arquivo, payload_completo

    except Exception as e:
        # Em caso de erro, lança uma exceção com mensagem personalizada
        raise RuntimeError(f"Erro ao gerar QR Code PIX: {str(e)}")


# =+-+=+-+=+-+=+-+=+-+=+-+= EMAIL =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/enviar-emaill', methods=['POST'])
def enviar_email():
    try:
        dados = request.json
        destinatario = dados.get('destinatario')

        if not destinatario:
            return jsonify({"erro": "Destinatário não fornecido"}), 400

        remetente = "bibliotecalibris@gmail.com"
        senha = "raij vzce iafk iekd"
        assunto = "Teste Simples"
        corpo = "Oi"

        msg = MIMEText(corpo, _charset='utf-8')
        msg['Subject'] = assunto
        msg['From'] = remetente
        msg['To'] = destinatario

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as servidor:
            servidor.login(remetente, senha)
            servidor.sendmail(remetente, [destinatario], msg.as_string())

        return jsonify({"mensagem": "E-mail enviado com sucesso!"}), 200

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


def enviar_email_com_template(destinatario, assunto, titulo, corpo, caminho_imagem=None):
    remetente = "bibliotecalibris@gmail.com"
    senha = "raij vzce iafk iekd"
    email.charset.add_charset('utf-8', email.charset.QP, email.charset.QP, 'utf-8')
    try:
        # Renderiza o HTML do template
        html_renderizado = render_template('template_email.html', titulo=titulo, corpo=corpo)
        # Cria a mensagem de e-mail
        msg = MIMEMultipart('alternative')
        # msg.attach(msg)

        # msg = MIMEMultipart('alternative')
        msg['Subject'] = assunto
        msg['From'] = remetente
        msg['To'] = destinatario
        msg.attach(MIMEText(html_renderizado, 'html', 'utf-8'))

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as servidor:
            servidor.login(remetente, senha)
            servidor.sendmail(remetente, [destinatario], msg.as_string())

        return True, "E-mail enviado com sucesso!"

        return False, f"Erro ao enviar e-mail: {e}"

    except Exception as e:
        return False, f"Erro ao enviar e-mail: {e}"


# =+-+=+-+=+-+=+-+=+-+=+-+= AVALIAÇÕES =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+

# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+ CRIAÇÃO DE AVALIAÇÕES =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/avaliacoes', methods=['POST'])
def criar_avaliacao():
    dados = request.get_json()

    id_usuario = dados.get('id_usuario')
    id_livro = dados.get('id_livro')
    voto = dados.get('voto')

    # Verificar se todos os campos foram fornecidos
    if not all([id_usuario, id_livro, voto]):
        return jsonify({"erro": "Campos obrigatórios: id_usuario, id_livro e voto"}), 400

    cursor = con.cursor()

    # Verificar se o usuário tem uma reserva válida
    cursor.execute("""
        SELECT 1 FROM RESERVAS 
        WHERE ID_USUARIO = ? AND ID_LIVRO = ? AND STATUS IN (2, 4)
    """, (id_usuario, id_livro))
    permissao = cursor.fetchone()

    if not permissao:
        cursor.close()
        return jsonify({"erro": "Você não leu esse livro, não pode avalia-lo"}), 403

    # Verificar se já existe uma avaliação do mesmo usuário para o mesmo livro
    cursor.execute("""
        SELECT 1 FROM AVALIACOES
        WHERE ID_USUARIO = ? AND ID_LIVRO = ?
    """, (id_usuario, id_livro))
    avaliacao_existente = cursor.fetchone()

    if avaliacao_existente:
        cursor.close()
        return jsonify({"erro": "Usuário já avaliou este livro."}), 409

    # Inserir a nova avaliação
    cursor.execute("""
        INSERT INTO AVALIACOES (ID_USUARIO, ID_LIVRO, VOTO)
        VALUES (?, ?, ?) RETURNING ID_AVALIACOES
    """, (id_usuario, id_livro, voto))

    id_avaliacao = cursor.fetchone()[0]
    con.commit()
    cursor.close()

    # Atualiza a média de avaliação no ACERVO
    atualizar_avaliacao_livro(id_livro)

    return jsonify({
        "mensagem": "Avaliação cadastrada com sucesso!",
        "avaliacao": {
            "id": id_avaliacao,
            "id_usuario": id_usuario,
            "id_livro": id_livro,
            "voto": voto
        }
    }), 201


# =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+ VIZUALIZAÇÃO DE AVALIAÇÕES =+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+=+-+
@app.route('/avaliacoes', methods=['GET'])
def obter_voto():
    id_usuario = request.args.get('id_usuario')
    id_livro = request.args.get('id_livro')

    # Verificar se os parâmetros foram fornecidos
    if not all([id_usuario, id_livro]):
        return jsonify({"erro": "Parâmetros obrigatórios: id_usuario e id_livro"}), 400

    cursor = con.cursor()
    cursor.execute("""
        SELECT VOTO FROM AVALIACOES
        WHERE ID_USUARIO = ? AND ID_LIVRO = ?
    """, (id_usuario, id_livro))

    resultado = cursor.fetchone()
    cursor.close()

    if resultado:
        return jsonify({
            "id_usuario": id_usuario,
            "id_livro": id_livro,
            "voto": resultado[0]
        }), 200
    else:
        return jsonify({"mensagem": "Nenhuma avaliação encontrada para este usuário e livro."}), 404


def atualizar_avaliacao_livro(id_livro):
    cursor = con.cursor()

    # Calcular a média dos votos do livro
    cursor.execute("""
        SELECT AVG(CAST(VOTO AS DECIMAL(10,2))) FROM AVALIACOES WHERE ID_LIVRO = ?;
    """, (id_livro,))
    media = cursor.fetchone()[0]

    if media is not None:
        # Atualizar a coluna AVALIACAO na tabela ACERVO
        cursor.execute("""
            UPDATE ACERVO SET AVALIACAO = ? WHERE ID_LIVRO = ?
        """, (media, id_livro))
        con.commit()

    cursor.close()
